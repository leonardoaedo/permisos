#!/usr/bin/python
# -*- encoding: utf-8 -*-
from xhtml2pdf import pisa
from easy_pdf.rendering import render_to_pdf_response
from easy_pdf.views import PDFTemplateView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import *
from cal.settings import *
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from django.contrib.staticfiles.templatetags.staticfiles import static
from django.views.generic import View
from reportlab.lib import colors
from reportlab.lib.pagesizes import *
from reportlab.platypus import *
from reportlab.lib.styles import *
from reportlab.lib.units import *
from django.db.models import Max
from dateutil.parser import *
from dateutil.tz import *
from django.utils.dateparse import parse_datetime
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, InvalidPage, EmptyPage
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.shortcuts import render
from django.shortcuts import redirect
from models import * 
from array import *
from numpy import *
from django.template import RequestContext, loader
from django.template.loader import render_to_string
from edt.models import *
from django.views.generic import ListView
from django.views.generic.dates import WeekArchiveView
from django.shortcuts import get_object_or_404,render_to_response
from django.forms import ModelForm
from django import forms
from calendar import month_name
from django.core.context_processors import csrf
from django.db.models import Count
from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.core.mail import EmailMessage
from icalendar import Calendar, Event
from pytz import timezone
from django.contrib.auth.decorators import login_required
from django.forms.models import modelformset_factory
from datetime import date, datetime, timedelta
from django.core import serializers
from funciones.session import estaLogeado
from funciones.funciones import *
from forms import *
from datetime import datetime
from dateutil import tz, parser 
from collections import Counter
from django.views.generic import ListView,DetailView
from django.views.generic.edit import CreateView,FormView,UpdateView
from django.views.generic.base import ContextMixin , TemplateView
from django.db.models import Count, Min, Sum, Avg , F
import calendar
import icalendar
import pytz
import os
import json
import urllib
import time
import tablib
from functools32 import lru_cache

# from edt.serializers import *

def ReemplazoAPI(request):
    if not estaLogeado(request):
        return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])
    lista = []

    if request.GET.get('licencia'):
        licencia = request.GET.get('licencia')

    reemplazos = ReemplazoLicencia.objects.all().order_by('-fecha').filter(licencia=licencia)

    for reemplazo in reemplazos:

        fecha = reemplazo.fecha.strftime('%Y-%m-%d')

        lista.append({ "licencia" : reemplazo.licencia.id,"reemplazante" : reemplazo.reemplazante.nombre,"horasreemplazo" : reemplazo.horasreemplazo,"horaspagar": reemplazo.horaspagar,"fecha" : fecha})
        data = json.dumps(lista)

    return HttpResponse(data,content_type='application/json')

def UsuarioAPI(request):
    if not estaLogeado(request):
        return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])

    lista = []

    licencias = Licencia.objects.all().values('funcionario').order_by("-inicio")

    funcionarios = Usuario.objects.all().filter(estado=1).filter(id__in=licencias)



    for funcionario in funcionarios :
        lista.append({"id":funcionario.id,"Nombre": funcionario.nombre,"Apellido1":funcionario.apellido1,"Apellido2":funcionario.apellido2})

        data = json.dumps(lista)

    return HttpResponse(data,content_type='application/json')

def LicenciaAPI(request):
    if not estaLogeado(request):
                return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])
    reemplazos = ReemplazoLicencia.objects.all().values('licencia')


    if request.GET.get('funcionario'):
        funcionario = request.GET.get('funcionario')

        lista  =[]
        licencias = Licencia.objects.all().filter(funcionario=funcionario)

        for licencia in licencias:

            inicio = licencia.inicio
            fin = licencia.fin
            inicio = inicio.strftime('%Y-%m-%d')
            fin = fin.strftime('%Y-%m-%d')
            funcionario = licencia.funcionario.id


            lista.append({"id":licencia.id,"inicio":inicio,"fin":fin,"funcionario" : funcionario})

            data = json.dumps(lista)

    return HttpResponse(data,content_type='application/json')


def rl(request):
    if not estaLogeado(request):
                return redirect("/login")
    hoy = datetime.now()
    user = Usuario.objects.get(id=request.session['usuario'])
    licencias = Licencia.objects.all().values('funcionario').order_by("-inicio")
    usuarios_filtro = Usuario.objects.all().filter(estado=1).filter(id__in=licencias)    
    reemplazantes_filtro = Usuario.objects.all().filter(estado=1)
    data = {}
    
    if request.POST:

        seleccionados = json.loads(request.POST['data-seleccionados'])

        funcionario = request.POST.get("funcionario")
        licencias_filtro = Licencia.objects.all().filter(funcionario_id=funcionario)
        for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(funcionario)

        if "licencia" in request.POST and request.POST.get("licencia") !="0":
            licencia = request.POST.get("licencia")                    
            for licen in licencias_filtro:
                licen.licencia_activa = licen.id == int(licencia)

        # if "horasreemplazo" in request.POST and request.POST.get("horasreemplazo") !=0:
        #             horasreemplazo = request.POST.get("horasreemplazo")

        # if "horaspagar" in request.POST and request.POST.get("horaspagar") >= 0:
        #             horaspagar = request.POST.get("horaspagar")    
                    
                
        # if "fecha" in request.POST and request.POST.get("fecha") !=" ":
        #             fecha = request.POST.get("fecha")

        for x in seleccionados:
            idreemplazante = x["idreemplazante"]
            idlicencia = x["idlicencia"]
            horaspagar = x["horaspagar"]
            horasreemplazo = x["horasreemplazo"]
            fecha = x["fecha"]
            idreemplazante = Usuario.objects.get(id=idreemplazante)
            idlicencia = Licencia.objects.get(id=idlicencia)
            reemplazolicencia = ReemplazoLicencia(horaspagar=horaspagar,fecha=fecha,licencia=idlicencia,reemplazante=idreemplazante,ingresadopor=user,horasreemplazo=horasreemplazo,fecha_ingreso=hoy)
            reemplazolicencia.save()



        return render_to_response("edt/guardareemplazo.html",data,context_instance=RequestContext(request))

        


        data = {
                "query_string" : request.META["QUERY_STRING"],
                "llego " : "llego un POST",
                "funcionario" : funcionario,
                "usuarios_filtro" :  usuarios_filtro,
                "licencias_filtro" : licencias_filtro,
                "usuario":user,
        }

        #return HttpResponse(json.dumps(seleccionados), content_type='application/json')

    if not estaLogeado(request):
        return redirect("/login")

    user = Usuario.objects.get(id=request.session['usuario'])
    licencias = Licencia.objects.all().values('funcionario').order_by("-inicio")
    funcionarios = Usuario.objects.all().filter(estado=1).filter(id__in=licencias)
    reemplazantes_filtro = Usuario.objects.all().filter(estado=1)
    data["funcionarios"] = funcionarios
    data["reemplazantes_filtro"] = reemplazantes_filtro
    data["usuario"] = user 


    return render_to_response("edt/rl.html",data,context_instance=RequestContext(request))

def InformeReemplazos(request):
    user = Usuario.objects.get(id=request.session['usuario'])
    licencias = Licencia.objects.all().values('funcionario').order_by("-inicio")
    funcionarios = Usuario.objects.all().filter(estado=1).filter(id__in=licencias)
    data = {
        "funcionarios" : funcionarios,
        "usuario" : user,
    }

    return render_to_response("edt/InformeReemplazos.html",data,context_instance=RequestContext(request))

# Generic View de Bitacora
class PermisoListView(ListView):

    template_name = 'edt/bitacora.html'
    context_object_name = 'bitacora'
    paginate_by = 10

    def get_queryset(self):
        return Bitacora.objects.order_by('-fecha')

#-----------------------------------------------------------------------------------------
def anulaciones(request):
    if not estaLogeado(request):
                return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])

    if  user.rol.id == 1:
        return HttpResponse ("Say Hiiiii!")
    else:
        return HttpResponse("Say Hooooo!")    



def login(request):
        if request.session.get('usuario',False):
                return redirect('/')#puse esa url para probar nomas
        else:
                if request.POST.get('username') and request.POST.get('password'):
                        usuario = request.POST.get('username')
                        password = request.POST.get('password')
                        try:
                                usuarioObj = Usuario.objects.get(username=usuario,password=password)
                                request.session['usuario'] = usuarioObj.id
                                return redirect('/main/')
                        except Exception as e:
                                return HttpResponseRedirect("/login/","El usuario o la contraseña son incorrectos!!!")


                else:
                        return render(request,'edt/login.html',{})

def logout(request):
            if not estaLogeado(request):
                return redirect("/login")

            try:
                    del request.session['usuario']
            except KeyError:
                    pass

            return redirect("/login")


def funcionarioEvento(request):
    if not estaLogeado(request):
        return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])

    if  user.rol.id == 1:
        eventos = Usuario.objects.annotate(cant_eventos=Count('edt_user')).order_by("cant_eventos").filter(estado=1)

   
    return render_to_response("edt/funcevento.html",{"eventos":eventos,"usuario": user},context_instance=RequestContext(request))    




def permiso_jefatura(request):

    if not estaLogeado(request):
                return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])
    formset = PermisoFormSet()
    formset.fields["reemplazante"].queryset = Usuario.objects.filter(estado=1)            

    if  user.rol.id == 1:
        usuarios_filtro = Usuario.objects.all().filter(estado=1)

        if "filtrar" in request.GET:
            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")                
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)    

        elif "limpiar" in request.GET:
                return redirect("/permiso_jefatura")


        data = {
                "form": formset,
                "usuarios_filtro" : usuarios_filtro,
                "usuario" : user,
                "persona" : request.GET.get("persona",None),
                "query_string" : request.META["QUERY_STRING"]           
        }

        return render_to_response("edt/permiso_jefatura.html",data,context_instance=RequestContext(request))
    else:
     return redirect("/main")         


def wsPermiso_Jefatura(request): # Web service que  genera calendario para cargar en pantalla 
        if not estaLogeado(request):
                return redirect("/login")
        user = Usuario.objects.get(id=request.session['usuario'])
        if  user.rol.id == 1:
       
            usuarios_filtro = Usuario.objects.all().filter(estado=1)

        if "filtrar" in request.GET:
            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")                
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)


        usuario_id=Usuario.objects.get(id=persona)
        #lista de los ids de los eventos ya ocupados
        hoy = datetime.now()
        fecha_cambio_TZ = parser.parse("Sep 04 2021 01:00AM")
        fecha_cambio_TZ2 = parser.parse("Apr 03 2022 01:00AM")
        inicio_agno = ("2019-02-15 00:00:00")
        fecha_inicio = parser.parse("Feb 15 2019 01:00AM")        
        fecha_inicio = 1000*(time.mktime(fecha_inicio.timetuple()))
        contador = 0


        ids = []
        for lista_ids in Eventos_en_Permisos.objects.all():
            ids.append(int(lista_ids.numero_evento.id))

         #data = serializers.serialize("json", Calendari.objects.filter(usuario_id=usuario_id))
         #Objeto.objects.all().exclude(id__in=lista) parta excluir los que estan en la lista
        lista = []
        for evento in Evento.objects.filter(usuario_id=usuario_id).filter(start__gte=inicio_agno):
                evento.flageado = evento.id in ids                   
                #formateo de timezone a milisegundos
                from_zone = tz.gettz('Europe/Paris') 

                start = datetime.timetuple(evento.start)                 
                start = time.strftime('%Y-%m-%dT%H:%M:%SZ', start)
                start = datetime.strptime(start, '%Y-%m-%dT%H:%M:%SZ')


                #automatizacion de cambio de  TZ a partir del 01 de noviembre
                if start >= fecha_cambio_TZ  and start <= fecha_cambio_TZ2:                   
                    to_zone = tz.gettz('America/Rio_Branco')# Regina -06 horas
                else:  
                    to_zone = tz.gettz('America/Rio_Branco') #  America/Rio_Branco  -05 ----------  America/Santo_Domingo : -04 desde 25-05 hasta ??? // resto del año America/Santiago - 03 // Noronha -02
                start = start.replace(tzinfo=from_zone)
                start = start.astimezone(to_zone)
                start = 1000*(time.mktime(start.timetuple()))
                end = datetime.timetuple(evento.end)
                end = time.strftime('%Y-%m-%dT%H:%M:%SZ', end)
                end = datetime.strptime(end, '%Y-%m-%dT%H:%M:%SZ')
                end = end.replace(tzinfo=from_zone)
                end = end.astimezone(to_zone)
                end = 1000 * (time.mktime(end.timetuple()))
                contador = contador + 1  
                #generacion de lista de eventos en formato json
                lista.append({"contador":contador,"fecha_inicio":fecha_inicio,"used": evento.flageado,"id":evento.id, "start":start, "end":end})
                data = json.dumps(lista)
            # "title":usuarioObj.nombre+" "+usuarioObj.apellido1 ,
                    
        #retorna la info en formato JSON
        return HttpResponse(data, content_type = "application/json")
        #return HttpResponse(json.dumps(ids)) ==> verificacion de permisos ya ocupados




def wsCalendario(request): # Web service que  genera calendario para cargar en pantalla 
        if not estaLogeado(request):
                return redirect("/login")
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])

        usuario_id=request.session['usuario']
        #lista de los ids de los eventos ya ocupados
        hoy = datetime.now()
        fecha_cambio_TZ = parser.parse("Sep 04 2021 01:00AM")
        fecha_cambio_TZ2 = parser.parse("Apr 03 2022 01:00AM")
        inicio_agno = ("2019-02-15 00:00:00")
        fecha_inicio = parser.parse("Feb 15 2019 01:00AM")        
        fecha_inicio = 1000*(time.mktime(fecha_inicio.timetuple()))
        contador = 0
        ids = []
        for lista_ids in Eventos_en_Permisos.objects.all():
            ids.append(int(lista_ids.numero_evento.id))
         #data = serializers.serialize("json", Calendari.objects.filter(usuario_id=usuario_id))
         #Objeto.objects.all().exclude(id__in=lista) parta excluir los que estan en la lista
        lista = []
        for evento in Evento.objects.filter(usuario_id=usuario_id).filter(start__gte=inicio_agno).iterator():
                evento.flageado = evento.id in ids                   
                #formateo de timezone a milisegundos
                from_zone = tz.gettz('Europe/Paris') 

                start = datetime.timetuple(evento.start)                 
                start = time.strftime('%Y-%m-%dT%H:%M:%SZ', start)
                start = datetime.strptime(start, '%Y-%m-%dT%H:%M:%SZ')


               #automatizacion de cambio de  TZ a partir del 01 de noviembre
                if start >= fecha_cambio_TZ  and start <= fecha_cambio_TZ2:                   
                    to_zone = tz.gettz('America/Rio_Branco')# Regina -06 horas
                else:  
                    to_zone = tz.gettz('America/Rio_Branco') # America/Santo_Domingo : -05 desde 25-05 hasta ??? // resto del año America/Santiago - 03 // Noronha -02
                
                start = start.replace(tzinfo=from_zone)
                start = start.astimezone(to_zone)
                start = 1000*(time.mktime(start.timetuple()))
                end = datetime.timetuple(evento.end)
                end = time.strftime('%Y-%m-%dT%H:%M:%SZ', end)
                end = datetime.strptime(end, '%Y-%m-%dT%H:%M:%SZ')
                end = end.replace(tzinfo=from_zone)
                end = end.astimezone(to_zone)
                end = 1000 * (time.mktime(end.timetuple()))
                contador = contador + 1  
                #generacion de lista de eventos en formato json
                lista.append({"contador":contador,"fecha_inicio":fecha_inicio,"used": evento.flageado,"id":evento.id, "start":start, "end":end})
                data = json.dumps(lista)
            # "title":usuarioObj.nombre+" "+usuarioObj.apellido1 ,
                    
        #retorna la info en formato JSON
        return HttpResponse(data, content_type = "application/json")
        #return HttpResponse(json.dumps(ids)) ==> verificacion de permisos ya ocupados

def comprobante(request, pk):
        if not estaLogeado(request):
            return redirect("/login")
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])
            
        #idpermiso = Permiso.objects.get(id=int(pk))        
        permiso = Permiso.objects.get(id=pk)
        #evento = Eventos_en_Permisos.objects.filter(numero_permiso=idpermiso)
	
	
 
        return render_to_response("edt/comprobante.html",{ "permiso" : permiso,"usuario" : usuarioObj})
        #return HttpResponse({ "permiso" : permiso,"usuario" : usuarioObj,"evento" : evento}) 

@csrf_exempt 
def urlcalendario(request): #procesa_solicitud_permiso
    if not estaLogeado(request):
                return redirect("/login")
    
    

    if request.POST.get('data-persona'):

        datapersona = request.POST.get('data-persona')        
        usuarioObj = Usuario.objects.get(id=datapersona)


    else:
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    actividad = Actividad.objects.get(id=5)
    usuario_id=request.session['usuario']


    if request.method == 'POST':
        arrayDeEventos = json.loads(request.POST['data-calendario'])  #cargo los id de los eventos seleccionados
        reemplazante = request.POST.get('reemplazante')
        devuelve_horas = request.POST.get('devuelve_horas')
        sueldo = request.POST.get('sueldo')
        comentario = request.POST.get('comentario')
        #comentario = comentario.decode('latin1','replace')# codificacion para tildes
        documento_adjunto = request.POST.get('documento_adjunto')
        #documento_adjunto = documento_adjunto.decode('latin1','replace')# codificacion para tildes
        motivo = request.POST.get('motivo')
                                                                      #en el calendario en formato json a una lista           
        eventos = Evento.objects.filter(id__in=arrayDeEventos) #consulto los eventos en la tabla Eventos segun los Id de la lista

        

    formset = PermisoFormSet(request.POST, request.FILES) 
   
    ############ #########guardado  de   permiso ############################

    if formset.is_valid():
        
        permiso = formset.save(commit=False)        
        permiso.usuario = usuarioObj
        reemplazante = reemplazante
        permiso.sueldo = sueldo
        permiso.comentario = comentario
        permiso.devuelve_horas = devuelve_horas
        permiso.documenot_adjunto = documento_adjunto        
        motivo = motivo
        permiso.save() # guardo los datos de Permiso en la BD

        if len(Permiso.objects.all()) == 0: # verifico si la tabla permisos esta vacia
            ultimopermiso = 1 # si esta vacia le asigno el valor 1
        else:
            #consulto el ultimo registro de la tabla Permiso
            ultimopermiso = Permiso.objects.all().order_by("-id")[0]

        suma = 0
        sumafuncionario = 0
        i = 0
        deltas = []
        deltaf = []  

        for evento in eventos:
            delta = evento.end - evento.start # calculo de la cantidad de horas solicitadas en segundos
            
            if (usuarioObj.estamento.id == 9 ):  
                #CALCULO PARA SECUNDARIA
                suma += float(delta.seconds) / 3300 # calculo para  Informes
                deltag = float(delta.seconds) / 3300 # calculo para  Informes
                deltas.append(round(deltag,2))
                sumafuncionario += float(delta.seconds) / 3300  # calculo para  funcionario
                deltafuncionario = float(delta.seconds) / 3300  # calculo para  funcionario
                deltaf.append(round(deltafuncionario,2))  
                  
                suma = round(suma,2) # redondeo a 2 decimales
                evento_en_permiso = Eventos_en_Permisos(numero_evento=evento,numero_permiso=ultimopermiso,deltainforme=deltas[i],deltafuncionario=deltaf[i])
                i += 1
                evento_en_permiso.save()


            if (usuarioObj.estamento.id == 5 or usuarioObj.estamento.id == 4 ):  
                #CALCULO PARA SECUNDARIA
                suma += float(delta.seconds) / 3300 # calculo para  Informes
                deltag = float(delta.seconds) / 3300 # calculo para  Informes
                deltas.append(round(deltag,2))
                sumafuncionario += float(delta.seconds) / 3300  # calculo para  funcionario
                deltafuncionario = float(delta.seconds) / 3300  # calculo para  funcionario
                deltaf.append(round(deltafuncionario,2))  
                  
                suma = round(suma,2) # redondeo a 2 decimales
                evento_en_permiso = Eventos_en_Permisos(numero_evento=evento,numero_permiso=ultimopermiso,deltainforme=deltas[i],deltafuncionario=deltaf[i])
                i += 1
                evento_en_permiso.save()
                

            if (usuarioObj.estamento.id == 2 or usuarioObj.estamento.id == 3):
                #CALCULO DE HORAS PARA PRIMARIA
                suma += float(delta.seconds) / 3600
                deltag = float(delta.seconds) / 3600
                deltas.append(round(deltag,2))
                suma = round(suma,2) # redondeo a 2 decimales
                sumafuncionario += float(delta.seconds) / 3600   # calculo para  funcionario
                deltafuncionario = float(delta.seconds) / 3600  # calculo para  funcionario
                deltaf.append(round(deltafuncionario,2))
                evento_en_permiso = Eventos_en_Permisos(numero_evento=evento,numero_permiso=ultimopermiso,deltainforme=deltas[i],deltafuncionario=deltaf[i])
                i += 1
                evento_en_permiso.save()
                        

            if (usuarioObj.estamento.id == 1 or usuarioObj.estamento.id == 6 or usuarioObj.estamento.id == 7 or usuarioObj.estamento.id == 8):
                #CALCULO DE HORAS PARA ADMINISTRACION,ASEM
                suma += float(delta.seconds) / 3600  
                deltag = float(delta.seconds) / 3600
                deltas.append(round(deltag,2))
                suma = round(suma,2) # redondeo a 2 decimales
                sumafuncionario = suma

                # sumafuncionario += float(delta.seconds) / 3600   # calculo para  funcionario
                # deltafuncionario = float(delta.seconds) / 3600  # calculo para  funcionario
                # deltaf.append(round(deltafuncionario,2))

                
                evento_en_permiso = Eventos_en_Permisos(numero_evento=evento,numero_permiso=ultimopermiso,deltainforme=deltas[i],deltafuncionario=deltas[i])
                i += 1
                evento_en_permiso.save()

        estado_permiso = Estado_Permiso.objects.get(id=3)
        permiso.estado = estado_permiso
        permiso.horas_solicitadas = suma
        permiso.horas_solicitadas_funcionario =  sumafuncionario
        permiso.save() # guardo suma aca despues de hacer el calculo   

        
        if usuarioObj.jefatura.id == 1 :
            folcpe = Foliocpe(permiso=ultimopermiso)
            folcpe.save()        

        if usuarioObj.jefatura.id == 4 :
            folprimaria = Folioprimaria(permiso=ultimopermiso)
            folprimaria.save()

        if usuarioObj.jefatura.id == 5 :
            folsecundaria = Foliosecundaria(permiso=ultimopermiso)
            folsecundaria.save()

        if usuarioObj.jefatura.id == 7 :
            folmantencion = Foliomantencion(permiso=ultimopermiso)
            folmantencion.save()

        if usuarioObj.jefatura.id == 3 :
            foldirgen = Foliodirgen(permiso=ultimopermiso)
            foldirgen.save()

        if usuarioObj.jefatura.id == 6 :
            folgerencia = Foliogerencia(permiso=ultimopermiso)
            folgerencia.save()


        ###################FOLIOS############################################################################

        cpe = Usuario.objects.values_list("jefatura").filter(jefatura__id=1)
        foliocpe = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=cpe))
    
        primaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=4)
        folioprimaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=primaria))

        secundaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=5)
        foliosecundaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=secundaria))

        mantencion = Usuario.objects.values_list("jefatura").filter(jefatura__id=7)
        foliomantencion = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=mantencion))

        dirgen = Usuario.objects.values_list("jefatura").filter(jefatura__id=3)
        foliosecundaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=dirgen))

        gerencia = Usuario.objects.values_list("jefatura").filter(jefatura__id=6)
        foliosecundaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=gerencia))

        ##########################################################################################################
        #guardo en horas para gestion de devolucion de horas
        
        horas = Horas(horas_solicitadas=suma,permiso=ultimopermiso,usuario=usuarioObj,horas_pendientes_por_aprobar=suma)
        horas.save()

        #guardado en bitacora
        bitacora = Bitacora(actividad=actividad,usuario=usuarioObj,permiso=ultimopermiso)            
        bitacora.save()


        #email a funcionario  idpermiso = Permiso.objects.get(pk=int(pk)) 
        template = loader.get_template('edt/email_solicitud.html')
        context = RequestContext(request, {'permiso' : ultimopermiso,'nombre' : usuarioObj.nombre,'apellido' : usuarioObj.apellido1,'horas':permiso.horas_solicitadas,'numero':permiso.id})
        html = template.render(context)
        msg = EmailMessage('Solicitud de permiso', html, 'scpa@cdegaulle.cl', [usuarioObj.correo])
        msg.content_subtype = "html"  # Main content is now text/html
        msg.send()

        #email a jefatura
        template = loader.get_template('edt/email_jefatura.html')
        context = RequestContext(request, {'permiso' : ultimopermiso,'jefe' : usuarioObj.jefatura.nombre,'horas':permiso.horas_solicitadas,'numero':permiso.id,'nombre' : usuarioObj.nombre,'apellido' : usuarioObj.apellido1,'rut': usuarioObj.rut,'dv':usuarioObj.dv,'nom_reemplazante':permiso.reemplazante.nombre,'ap_reemplazante':permiso.reemplazante.apellido1,'fecha': permiso.fecha_creacion})
        html = template.render(context)
        msg = EmailMessage('Solicitud de permiso ' , html, 'scpa@cdegaulle.cl', [usuarioObj.jefatura.correo1])
        msg.content_subtype = "html"  # Main content is now text/html
        msg.send()

        
        return redirect("/comprobante/%d"%(ultimopermiso.id))
        #return HttpResponse(sumafuncionario)
    # data = json.dumps({"suma" : suma,"ultimo" : ultimopermiso,"largo":len(eventos),"usuario":usuario_id,"reemplazante" : reemplazante,"devuelve_horas":devuelve_horas})
    # return HttpResponse(data, content_type = "application/json")
    else:
    #    formset = PermisoFormSet()
        data = {
            "form": formset,
            "usuario": usuarioObj,
            #"fech" : fech,
           }

        return render_to_response("edt/main.html",data,context_instance=RequestContext(request))


def bitgeneral(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    
    if  usuarioObj.rol.id == 1:
        #permiso = Permiso.objects.all()
        permisos = Permiso.objects.all()
        #resolucion =  Resolucion.objects.all().order_by('-id')
        return render_to_response("edt/bgeneral.html",{"permisos" : permisos,"usuario": usuarioObj},context_instance=RequestContext(request))
    else:
     return redirect("/main")

def EstadisticaPermisos(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    if usuarioObj.rol.id == 1:

        permisos = Permiso.objects.all().order_by("usuario__apellido1").filter(usuario__estado=1)
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()
        thora = "Seleccione Opción";
         
        if "filtrar" in request.GET:
            if "start" in request.GET and request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=request.GET.get("start"))
                start = request.GET.get("start")


            if "end" in request.GET and request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=request.GET.get("end"))
                end = request.GET.get("end")


            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                permisos = permisos.filter(usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

            if "thora" in request.GET and request.GET.get("thora") != "0" :
                thora = request.GET.get("thora")
                if thora == 'Horas por Descontar':
                    permisos = permisos.filter(horas__horas_descontar__gt=0)
                if thora == 'Horas por Devolver':
                    permisos = permisos.filter(horas__horas_por_devolver__gt=0)
                if thora == 'Horas Descontadas':
                    permisos = permisos.filter(horas__horas_descontadas__gt=0)
                if thora == 'Horas Devueltas':
                    permisos = permisos.filter(horas__horas_devueltas__gt=0)

        elif "limpiar" in request.GET:
                return redirect("/estadisticapermisos")
            
        data = {
                "usuario": usuarioObj,
                "permisos" : permisos,
                "query_string" : request.META["QUERY_STRING"],
                "usuarios_filtro" : usuarios_filtro,
                "estamento_filtro" : estamento_filtro,
               }


        #Filtros para generar la fecha e el Titulo
        if "filtrar" in request.GET:
            if "start" in request.GET and request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=request.GET.get("start"))
                start = request.GET.get("start")
                inicio = datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
                data["inicio"] = inicio
                data["start"] = start 


            if "end" in request.GET and request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=request.GET.get("end"))
                end = request.GET.get("end")
                termino = datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
                data["termino"] = termino
                data["end"] = end       

        return render_to_response("edt/estadisticapermisos.html",data,context_instance=RequestContext(request))  

    else:
        return redirect("/main")    


def bithoras(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    if usuarioObj.rol.id == 1:

        permisos = Permiso.objects.all().order_by("usuario__apellido1").filter(usuario__estado=1)
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()
        thora = "Seleccione Opción";
         
        if "filtrar" in request.GET:
            if "start" in request.GET and request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=request.GET.get("start"))
                start = request.GET.get("start")


            if "end" in request.GET and request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=request.GET.get("end"))
                end = request.GET.get("end")


            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                permisos = permisos.filter(usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

            if "thora" in request.GET and request.GET.get("thora") != "0" :
                thora = request.GET.get("thora")
                if thora == 'Horas por Descontar':
                    permisos = permisos.filter(horas__horas_descontar__gt=0)
                if thora == 'Horas por Devolver':
                    permisos = permisos.filter(horas__horas_por_devolver__gt=0)
                if thora == 'Horas Descontadas':
                    permisos = permisos.filter(horas__horas_descontadas__gt=0)
                if thora == 'Horas Devueltas':
                    permisos = permisos.filter(horas__horas_devueltas__gt=0)


        elif "limpiar" in request.GET:
            return redirect("/bhoras")

        #else :
         #   return render_to_response("edt/bhoras.html",data,context_instance=RequestContext(request))

        usuarios = {}
        for permiso in permisos:
            idUsuario = permiso.usuario.id
            if idUsuario not in usuarios:
                usuarios[idUsuario] = {
                     "id" : permiso.usuario.id,
                     "nombre" : permiso.usuario.nombre,
                     "apellido1"  :permiso.usuario.apellido1,
                     "apellido2" : permiso.usuario.apellido2,
                     "estamento" : permiso.usuario.estamento.nombre,
                     "total_horas" : 0,
                     "aprobadas" : 0,
                     "rechazadas" : 0,
                     "devolveracumuladas" : 0,
                     "devueltas" : 0,
                     "saldodevolucion" : 0,
                     "descontaracumuladas" : 0,
                     "descontadas" : 0,
                     "saldodescontar" : 0,
                     "sin_recup_con_sueldo" : 0,
                     "pendientes_por_aprobar" : 0,
                 }
             
            horas = permiso.horas_set.all()
            #traspaso de datos a otra columna
            # for hora in horas :
            #     hora.horas_por_devolver_acumuladas = hora.horas_por_devolver
            #     hora.horas_descontar_acumuladas = hora.horas_descontar
            #     hora.save()
 
            if len(horas) == 0:
                 continue
 
            hora = horas[0]
            usuarios[idUsuario]["total_horas"] += hora.horas_solicitadas
            usuarios[idUsuario]["aprobadas"] += hora.horas_aprobadas
            usuarios[idUsuario]["rechazadas"] += hora.horas_rechazadas
            usuarios[idUsuario]["devolveracumuladas"] += hora.horas_por_devolver_acumuladas
            usuarios[idUsuario]["devueltas"] += hora.horas_devueltas
            usuarios[idUsuario]["saldodevolucion"] += hora.horas_por_devolver
            usuarios[idUsuario]["descontaracumuladas"] += hora.horas_descontar_acumuladas
            usuarios[idUsuario]["descontadas"] += hora.horas_descontadas
            usuarios[idUsuario]["saldodescontar"] += hora.horas_descontar
            usuarios[idUsuario]["sin_recup_con_sueldo"] += hora.horas_sin_recuperacion_con_goce
            usuarios[idUsuario]["pendientes_por_aprobar"] += hora.horas_pendientes_por_aprobar
             
        usuariosLista = [value for key,value in usuarios.iteritems()]
 
        # headers = ('Nombre','Estamento','Horas solicitadas')
        # excel = []
        # excel = tablib.Dataset()
        # excel.headers = 

        data = {

            "usuario": usuarioObj,
            "usuarios" : usuariosLista,
            "usuarios_filtro" : usuarios_filtro,
            "estamento_filtro" : estamento_filtro,
            "thora" :thora,
            "query_string" : request.META["QUERY_STRING"]
                }

        #Filtros para generar la fecha e el Titulo
        if "filtrar" in request.GET:
            if "start" in request.GET and request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=request.GET.get("start"))
                start = request.GET.get("start")
                inicio = datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
                data["inicio"] = inicio
                data["start"] = start 


            if "end" in request.GET and request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=request.GET.get("end"))
                end = request.GET.get("end")
                termino = datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
                data["termino"] = termino
                data["end"] = end



        return render_to_response("edt/bhoras.html",data,context_instance=RequestContext(request))

        
        #return HttpResponse(json.dumps(usuarios)) 
        #return HttpResponse(json.dumps(usuariosLista))

    else:
        return redirect("/main")

##########Generacion de informes################

class EstadisticaPermisosExcel(TemplateView):
    """docstring for BitHorasExcel"""
    def get(self, request, *args, **kwargs):

        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        permisos = Permiso.objects.all().order_by("id").filter(usuario__estado=1)
        eventos_en_permisos = Eventos_en_Permisos.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()
        thora = "Seleccione Opción";
         
        if "filtrar" in self.request.GET:
            if "start" in self.request.GET and self.request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=self.request.GET.get("start"))
                start = self.request.GET.get("start")


            if "end" in self.request.GET and self.request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=self.request.GET.get("end"))
                end = self.request.GET.get("end")

        if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                permisos = permisos.filter(usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

        if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        if "thora" in self.request.GET and self.request.GET.get("thora") != "0" :
                thora = self.request.GET.get("thora")
                if thora == 'Horas por Descontar':
                    permisos = permisos.filter(horas__horas_descontar__gt=0)
                if thora == 'Horas por Devolver':
                    permisos = permisos.filter(horas__horas_por_devolver__gt=0)
                if thora == 'Horas Descontadas':
                    permisos = permisos.filter(horas__horas_descontadas__gt=0)
                if thora == 'Horas Devueltas':
                    permisos = permisos.filter(horas__horas_devueltas__gt=0)        


        hoy = datetime.now()
        #Creamos el libro de trabajo
        wb= Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Estadistica Permisos"
        ws.sheet_properties.tabColor = "1072BA"
        
        # border = Border(left=Side(border_style='medium'),
        #                  right=Side(border_style='medium'),
        #                  top=Side(border_style='medium'),
        #                  bottom=Side(border_style='medium'))
        # ws.border = border
        #En la celda B1 ponemos el texto 'ESTADISTICA PERMISOS'
        ws['B1'] = 'ESTADISTICA PERMISOS'
        ws['B2'] = hoy
        #definimos color de letra para los eventos anulados
        letra_roja = Font(color="FF0101")
        letra_azul = Font(color="1D31E6")
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la L3
        ws['B3'] = 'Correlativo'
        ws['C3'] = 'Solicitante'
        ws['D3'] = 'Rut'
        ws['E3'] = 'Funcion/Cargo'
        ws['F3'] = 'Fecha de solicitud'
        ws['G3'] = 'Tipo'
        ws['H3'] = 'Dia del permiso'
        ws['I3'] = 'Motivo'
        ws['J3'] = 'Comentario'
        ws['K3'] = 'Horas'
        ws['L3'] = 'Bloques'
        ws['M3'] = 'Resolución'
        ws['N3'] = 'Devuelve horas'
        ws['O3'] = 'Con o sin goce de sueldo'
        ws['P3'] = 'Reemplazante'        
        ws['Q3'] = 'Archivo adjunto'


        cont=4
        for permiso in permisos:

            if permiso.estado:
                if permiso.estado.id == 6: 
                    ws.cell(row=cont,column=2).font = letra_roja
                    ws.cell(row=cont,column=3).font = letra_roja
                    ws.cell(row=cont,column=4).font = letra_roja
                    ws.cell(row=cont,column=5).font = letra_roja
                    ws.cell(row=cont,column=6).font = letra_roja
                    ws.cell(row=cont,column=7).font = letra_roja
                    ws.cell(row=cont,column=8).font = letra_roja
                    ws.cell(row=cont,column=9).font = letra_roja
                    ws.cell(row=cont,column=10).font = letra_roja
                    ws.cell(row=cont,column=11).font = letra_roja
                    ws.cell(row=cont,column=12).font = letra_roja
                    ws.cell(row=cont,column=13).font = letra_roja
                    ws.cell(row=cont,column=14).font = letra_roja
                    ws.cell(row=cont,column=15).font = letra_roja
                    ws.cell(row=cont,column=16).font = letra_roja
                    ws.cell(row=cont,column=17).font = letra_roja

            ws.cell(row=cont,column=2).value = permiso.id
            ws.cell(row=cont,column=3).value = permiso.usuario.apellido1+" "+permiso.usuario.apellido2+" "+permiso.usuario.nombre
            ws.cell(row=cont,column=4).value = permiso.usuario.rut+"-"+permiso.usuario.dv
            ws.cell(row=cont,column=5).value = permiso.usuario.cargo.nombre
            ws.cell(row=cont,column=6).value = permiso.fecha_creacion
            ws.cell(row=cont,column=7).value = permiso.tipo.nombre
            if permiso.primerEvento():
                ws.cell(row=cont,column=8).value = permiso.primerEvento().numero_evento.start
            ws.cell(row=cont,column=9).value = permiso.motivo.nombre
            ws.cell(row=cont,column=10).value = permiso.comentario
            if permiso.usuario.cargo.id != 19:  
                ws.cell(row=cont,column=11).value = float(permiso.horas_solicitadas)    
            if permiso.usuario.cargo.id == 19:  
                ws.cell(row=cont,column=12).value = float(permiso.horas_solicitadas)            
            if permiso.ultimaResolucion():
                ws.cell(row=cont,column=13).value = permiso.ultimaResolucion().get_respuesta_display()
            else :
                ws.cell(row=cont,column=13).value = 'Sin Revisar'
                ws.cell(row=cont,column=13).font = letra_azul    
            ws.cell(row=cont,column=14).value = permiso.get_devuelve_horas_display()
            ws.cell(row=cont,column=15).value = permiso.get_sueldo_display()
            ws.cell(row=cont,column=16).value = permiso.reemplazante.apellido1+" "+permiso.reemplazante.nombre
            if permiso.documento_adjunto.name:
                ws.cell(row=cont,column=17).value =  "Con documento Adjunto"      
            cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="estadistica_pemisos.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class EstadisticaPermisosPDF(PDFTemplateView):
    template_name = "edt/estadisticapermisosPDF.html"

    def get_context_data(self, **kwargs):
        context =  super(EstadisticaPermisosPDF,self).get_context_data(
            pagesize="letter",
            title="Informe PDF",
            **kwargs
            )

        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        permisos = Permiso.objects.all().order_by("id").filter(usuario__estado=1)
        eventos_en_permisos = Eventos_en_Permisos.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()
        thora = "Seleccione Opción";

        if "filtrar" in self.request.GET:
            if "start" in self.request.GET and self.request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=self.request.GET.get("start"))
                start = self.request.GET.get("start")


            if "end" in self.request.GET and self.request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=self.request.GET.get("end"))
                end = self.request.GET.get("end")

        if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                permisos = permisos.filter(usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

        if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        if "thora" in self.request.GET and self.request.GET.get("thora") != "0" :
                thora = self.request.GET.get("thora")
                if thora == 'Horas por Descontar':
                    permisos = permisos.filter(horas__horas_descontar__gt=0)
                if thora == 'Horas por Devolver':
                    permisos = permisos.filter(horas__horas_por_devolver__gt=0)
                if thora == 'Horas Descontadas':
                    permisos = permisos.filter(horas__horas_descontadas__gt=0)
                if thora == 'Horas Devueltas':
                    permisos = permisos.filter(horas__horas_devueltas__gt=0)


        context = {
                "usuario": usuarioObj,
                "permisos" : permisos,
                "usuarios_filtro" : usuarios_filtro,
                "estamento_filtro" : estamento_filtro,
                "query_string" : self.request.META["QUERY_STRING"],
               }
        context['hoy'] =  datetime.now()       


        #Filtros para generar la fecha e el Titulo
        if "filtrar" in self.request.GET:
            if "start" in self.request.GET and self.request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=self.request.GET.get("start"))
                start = self.request.GET.get("start")
                inicio = datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
                context["inicio"] = inicio
                context["start"] = start 


            if "end" in self.request.GET and self.request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=self.request.GET.get("end"))
                end = self.request.GET.get("end")
                termino = datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
                context["termino"] = termino
                context["end"] = end

        return context


class BitHorasPDF(PDFTemplateView):
    template_name= "edt/bhorasPDF.html"

    def get_context_data(self, **kwargs):
        context =  super(BitHorasPDF,self).get_context_data(
            pagesize="letter",
            title="Informe PDF",
            **kwargs
            )

        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        permisos = Permiso.objects.all().order_by("usuario__apellido1").filter(usuario__estado=1)
        estamento = Estamento.objects.all() 
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()
        thora = "Seleccione Opción";
        
         
        if "filtrar" in self.request.GET:
            if "start" in self.request.GET and self.request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=self.request.GET.get("start"))
                start = self.request.GET.get("start")


            if "end" in self.request.GET and self.request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=self.request.GET.get("end"))
                end = self.request.GET.get("end")


            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                permisos = permisos.filter(usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

            if "thora" in self.request.GET and self.request.GET.get("thora") != "0" :
                thora = self.request.GET.get("thora")
                if thora == 'Horas por Descontar':
                    permisos = permisos.filter(horas__horas_descontar__gt=0)
                if thora == 'Horas por Devolver':
                    permisos = permisos.filter(horas__horas_por_devolver__gt=0)
                if thora == 'Horas Descontadas':
                    permisos = permisos.filter(horas__horas_descontadas__gt=0)
                if thora == 'Horas Devueltas':
                    permisos = permisos.filter(horas__horas_devueltas__gt=0)

        usuarios = {}
        for permiso in permisos:
            idUsuario = permiso.usuario.id
            if idUsuario not in usuarios:
                usuarios[idUsuario] = {
                     "id" : permiso.usuario.id,
                     "nombre" : permiso.usuario.nombre,
                     "apellido1"  :permiso.usuario.apellido1,
                     "apellido2" : permiso.usuario.apellido2,
                     "estamento" : permiso.usuario.estamento.nombre,
                     "total_horas" : 0,
                     "aprobadas" : 0,
                     "rechazadas" : 0,
                     "devolveracumuladas" : 0,
                     "devueltas" : 0,
                     "saldodevolucion" : 0,
                     "descontaracumuladas" : 0,
                     "descontadas" : 0,
                     "saldodescontar" : 0,
                     "sin_recup_con_sueldo" : 0,
                     "pendientes_por_aprobar" : 0,
                 }
             
            horas = permiso.horas_set.all()
            #traspaso de datos a otra columna
            # for hora in horas :
            #     hora.horas_por_devolver_acumuladas = hora.horas_por_devolver
            #     hora.horas_descontar_acumuladas = hora.horas_descontar
            #     hora.save()
 
            if len(horas) == 0:
                 continue
 
            hora = horas[0]
            usuarios[idUsuario]["total_horas"] += hora.horas_solicitadas
            usuarios[idUsuario]["aprobadas"] += hora.horas_aprobadas
            usuarios[idUsuario]["rechazadas"] += hora.horas_rechazadas
            usuarios[idUsuario]["devolveracumuladas"] += hora.horas_por_devolver_acumuladas
            usuarios[idUsuario]["devueltas"] += hora.horas_devueltas
            usuarios[idUsuario]["saldodevolucion"] += hora.horas_por_devolver
            usuarios[idUsuario]["descontaracumuladas"] += hora.horas_descontar_acumuladas
            usuarios[idUsuario]["descontadas"] += hora.horas_descontadas
            usuarios[idUsuario]["saldodescontar"] += hora.horas_descontar
            usuarios[idUsuario]["sin_recup_con_sueldo"] += hora.horas_sin_recuperacion_con_goce
            usuarios[idUsuario]["pendientes_por_aprobar"] += hora.horas_pendientes_por_aprobar
             
        usuariosLista = [value for key,value in usuarios.iteritems()]
 
      

        context = {

            "usuario": usuarioObj,
            "usuarios" : usuariosLista,
            "usuarios_filtro" : usuarios_filtro,
            "estamento_filtro" : estamento_filtro,
            "thora" :thora,            
                }
        context['hoy'] =  datetime.now()
        #Filtros para generar la fecha e el Titulo
        
        if "filtrar" in self.request.GET:
            if "start" in self.request.GET and self.request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=self.request.GET.get("start"))
                start = self.request.GET.get("start")
                inicio = datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
                context["inicio"] = inicio
                context["start"] = start 


            if "end" in self.request.GET and self.request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=self.request.GET.get("end"))
                end = self.request.GET.get("end")
                termino = datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
                context["termino"] = termino
                context["end"] = end
        
        return context


class BitHorasExcel(TemplateView):
    """docstring for BitHorasExcel"""
    def get(self, request, *args, **kwargs):

        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        permisos = Permiso.objects.all().order_by("usuario__apellido1").filter(usuario__estado=1)
        estamento = Estamento.objects.all() 
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()
        thora = "Seleccione Opción";
        
         
        if "filtrar" in self.request.GET:
            if "start" in self.request.GET and self.request.GET.get("start") != "":
                permisos = permisos.filter(fecha_creacion__gte=self.request.GET.get("start"))
                start = self.request.GET.get("start")


            if "end" in self.request.GET and self.request.GET.get("end") != "":
                permisos = permisos.filter(fecha_creacion__lte=self.request.GET.get("end"))
                end = self.request.GET.get("end")


            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                permisos = permisos.filter(usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

            if "thora" in self.request.GET and self.request.GET.get("thora") != "0" :
                thora = self.request.GET.get("thora")
                if thora == 'Horas por Descontar':
                    permisos = permisos.filter(horas__horas_descontar__gt=0)
                if thora == 'Horas por Devolver':
                    permisos = permisos.filter(horas__horas_por_devolver__gt=0)
                if thora == 'Horas Descontadas':
                    permisos = permisos.filter(horas__horas_descontadas__gt=0)
                if thora == 'Horas Devueltas':
                    permisos = permisos.filter(horas__horas_devueltas__gt=0)

        usuarios = {}
        for permiso in permisos:
            idUsuario = permiso.usuario.id
            if idUsuario not in usuarios:
                usuarios[idUsuario] = {
                     "id" : permiso.usuario.id,
                     "nombre" : permiso.usuario.nombre,
                     "apellido1"  :permiso.usuario.apellido1,
                     "apellido2" : permiso.usuario.apellido2,
                     "estamento" : permiso.usuario.estamento.nombre,
                     "total_horas" : 0,
                     "aprobadas" : 0,
                     "rechazadas" : 0,
                     "devolveracumuladas" : 0,
                     "devueltas" : 0,
                     "saldodevolucion" : 0,
                     "descontaracumuladas" : 0,
                     "descontadas" : 0,
                     "saldodescontar" : 0,
                     "sin_recup_con_sueldo" : 0,
                     "pendientes_por_aprobar" : 0,
                 }
             
            horas = permiso.horas_set.all()
            #traspaso de datos a otra columna
            # for hora in horas :
            #     hora.horas_por_devolver_acumuladas = hora.horas_por_devolver
            #     hora.horas_descontar_acumuladas = hora.horas_descontar
            #     hora.save()
 
            if len(horas) == 0:
                 continue
 
            hora = horas[0]
            usuarios[idUsuario]["total_horas"] += hora.horas_solicitadas
            usuarios[idUsuario]["aprobadas"] += hora.horas_aprobadas
            usuarios[idUsuario]["rechazadas"] += hora.horas_rechazadas
            usuarios[idUsuario]["devolveracumuladas"] += hora.horas_por_devolver_acumuladas
            usuarios[idUsuario]["devueltas"] += hora.horas_devueltas
            usuarios[idUsuario]["saldodevolucion"] += hora.horas_por_devolver
            usuarios[idUsuario]["descontaracumuladas"] += hora.horas_descontar_acumuladas
            usuarios[idUsuario]["descontadas"] += hora.horas_descontadas
            usuarios[idUsuario]["saldodescontar"] += hora.horas_descontar
            usuarios[idUsuario]["sin_recup_con_sueldo"] += hora.horas_sin_recuperacion_con_goce
            usuarios[idUsuario]["pendientes_por_aprobar"] += hora.horas_pendientes_por_aprobar
             
             
        usuariosLista = [value for key,value in usuarios.iteritems()]

        hoy = datetime.now()
        #Creamos el libro de trabajo
        wb= Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'INFORME DE HORAS'
        ws['B1'] = 'INFORME DE HORAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la L3
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'ESTAMENTO'
        ws['D3'] = 'HORAS SOLICITADAS'
        ws['E3'] = 'HORAS APROBADAS'
        ws['F3'] = 'HORAS RECHAZADAS'
        ws['G3'] = 'HORAS A RECUPERAR ACUMULADAS'
        ws['H3'] = 'HORAS RECUPERADAS'
        ws['I3'] = 'SALDO DE HORAS POR RECUPERAR'
        ws['J3'] = 'HORAS POR DESCONTAR ACUMULADAS'
        ws['K3'] = 'HORAS DESCONTADAS'
        ws['L3'] = 'SALDO HORAS POR DESCONTAR'
        ws['M3'] = 'HORAS SIN RECUPERACION CON GOCE DE SUELDO'
        ws['N3'] = 'HORAS PENDIENTES DE REVISION'
        cont=4
        for usuario in usuariosLista:

            ws.cell(row=cont,column=2).value = usuario["apellido1"]+" "+usuario["apellido2"]+" "+usuario["nombre"]
            ws.cell(row=cont,column=3).value = usuario["estamento"]
            ws.cell(row=cont,column=4).value = usuario["total_horas"]
            ws.cell(row=cont,column=5).value = usuario["aprobadas"]
            ws.cell(row=cont,column=6).value = usuario["rechazadas"]
            ws.cell(row=cont,column=7).value = usuario["devolveracumuladas"]
            ws.cell(row=cont,column=8).value = usuario["devueltas"]
            ws.cell(row=cont,column=9).value = usuario["saldodevolucion"]  
            ws.cell(row=cont,column=10).value = usuario["descontaracumuladas"]
            ws.cell(row=cont,column=11).value = usuario["descontadas"]
            ws.cell(row=cont,column=12).value = usuario["saldodescontar"]
            ws.cell(row=cont,column=13).value = usuario["sin_recup_con_sueldo"]
            ws.cell(row=cont,column=14).value = usuario["pendientes_por_aprobar"]


            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="informe_de_horas.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response     
        



##############FIN GENERACION DE INFORMES ###################################
def permisosusuario(request,pk):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    permisos = Permiso.objects.filter(usuario__id=pk)
    bitacoras = Bitacora.objects.filter(permiso__usuario__id=pk)
    horas = Horas.objects.filter(permiso__usuario__id=pk)

    data = {

        "usuario" : usuarioObj,
        "permisos" : permisos,
        "months" : mkmonth_lst(),   
        "bitacoras" : bitacoras,
        "horas" : horas,

    }


    return render_to_response("edt/permisosusuariolst.html",data,context_instance=RequestContext(request))




def bitfuncionario(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    
    #if  usuarioObj.rol.id == 1:
        
    permisos = Permiso.objects.filter(usuario=usuarioObj.id).annotate()
    #resolucion =  Resolucion.objects.all().order_by('-id')
    return render_to_response("edt/bfuncionario.html",{"permisos" : permisos,"usuario": usuarioObj},context_instance=RequestContext(request))
    # else:
    #  return redirect("/main")        

def index(request):
        if not estaLogeado(request):
                return redirect("/login")
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])
	
        if  usuarioObj.rol.id == 1:

            if request.method == 'POST':
                formset = DocumentFormSet(request.POST, request.FILES)
                if formset.is_valid():
                        document= formset.save(commit=False)
                        document.save()
                        return redirect('/upload/%d'%(document.id))
            
                
            formset = DocumentFormSet()
            
            return render_to_response("edt/index.html", {"form": formset,"usuario": usuarioObj},context_instance=RequestContext(request))
        else:
                return redirect("/main") 

def permisolst(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
    rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')
    gerencia = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__gte=1).filter(usuario__estamento=1)
    dirgen = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__gte=2).filter(usuario__jefatura=3)
    revisado_gerente = Resolucion.objects.values_list("permiso").filter(resolutor=39)

    permisos = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__lte=1).exclude(id__in=revisado_gerente).exclude(id__in=dirgen).exclude(id__in=gerencia).exclude(id__in=anulados).exclude(id__in=rechazados).order_by("-fecha_creacion")
    

    bitacoras = Bitacora.objects.all()
    revisores = Revisor.objects.all()


    cpe = Usuario.objects.values_list("jefatura").filter(jefatura__id=1).filter(estado=1)
    foliocpe = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=cpe))    
    primaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=4).filter(estado=1)
    folioprimaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=primaria))
    secundaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=5).filter(estado=1)
    foliosecundaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=secundaria))    

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                permisos = permisos.filter(usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/permisolst")       
        
        paginator = Paginator(permisos,30)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            permisos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            permisos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,
                 "foliocpe":foliocpe,
                 "folioprimaria" :folioprimaria,
                 "foliosecundaria" :foliosecundaria ,
                 "permisos": permisos,
                 "usuario" : usuarioObj,
                 "permisos_list" : permisos.object_list,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "bitacoras" : bitacoras,
                 "revisores" : revisores,
                 }


        return render_to_response("edt/permisolst.html",data)
        #return HttpResponse(administracion)

    else:
        #if     usuarioObj.username == request.session['usuario']:
        if  usuarioObj.rol.nivel_acceso == 1:
            ids = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
            permiso = Permiso.objects.annotate(num_b=Count('resolucion')).filter(usuario=usuarioObj.id).exclude(id__in=ids).order_by("-fecha_creacion")
            

            paginator = Paginator(permiso,10)       
            try: pagina = int(request.GET.get("page",'1'))
            except ValueError: pagina = 1       
            try:
                permiso = paginator.page(pagina)
            except (InvalidPage, EmptyPage):
                permiso = paginator.page(paginator.num_pages)

        data = {
                 "foliocpe":foliocpe,
                 "folioprimaria" :folioprimaria,
                 "foliosecundaria" :foliosecundaria ,
                 "permisos": permiso,
                 "usuario" : usuarioObj,
                 "permisos_list" : permiso.object_list,
                 "months" : mkmonth_lst(),
               }
            
        return render_to_response("edt/permisolst.html",data)


def modpermisolst(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
    rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')
    gerencia = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__gte=1).filter(usuario__estamento=1)
    dirgen = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__gte=2).filter(usuario__jefatura=3)

    permisos = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__gte=2).exclude(id__in=anulados).exclude(id__in=rechazados).order_by("-fecha_creacion") 
    #.exclude(id__in=dirgen).exclude(id__in=gerencia)

    bitacoras = Bitacora.objects.all()
    revisores = Revisor.objects.all()


    cpe = Usuario.objects.values_list("jefatura").filter(jefatura__id=1).filter(estado=1)
    foliocpe = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=cpe))    
    primaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=4).filter(estado=1)
    folioprimaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=primaria))
    secundaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=5).filter(estado=1)
    foliosecundaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=secundaria))    

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                permisos = permisos.filter(usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/modpermisolst")       
        
        paginator = Paginator(permisos,30)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            permisos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            permisos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,
                 "foliocpe":foliocpe,
                 "folioprimaria" :folioprimaria,
                 "foliosecundaria" :foliosecundaria ,
                 "permisos": permisos,
                 "usuario" : usuarioObj,
                 "permisos_list" : permisos.object_list,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "bitacoras" : bitacoras,
                 "revisores" : revisores,
                 }


        return render_to_response("edt/modpermisolst.html",data)
        #return HttpResponse(administracion)

    else:
        #if     usuarioObj.username == request.session['usuario']:
        if  usuarioObj.rol.nivel_acceso == 1:
            ids = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
            permiso = Permiso.objects.annotate(num_b=Count('resolucion')).filter(usuario=usuarioObj.id).exclude(id__in=ids).order_by("-fecha_creacion")
            

            paginator = Paginator(permiso,10)       
            try: pagina = int(request.GET.get("page",'1'))
            except ValueError: pagina = 1       
            try:
                permiso = paginator.page(pagina)
            except (InvalidPage, EmptyPage):
                permiso = paginator.page(paginator.num_pages)

        data = {
                 "foliocpe":foliocpe,
                 "folioprimaria" :folioprimaria,
                 "foliosecundaria" :foliosecundaria ,
                 "permisos": permiso,
                 "usuario" : usuarioObj,
                 "permisos_list" : permiso.object_list,
                 "months" : mkmonth_lst(),
               }
            
        return render_to_response("edt/modpermisolst.html",data)

def modpermiso(request, pk):
     if not estaLogeado(request):
            return redirect("/login")
    
     permiso = Permiso.objects.get(id=pk)
     permiso_formset = PermisoFormSet(instance=permiso)
     usuarioObj = Usuario.objects.get(id=request.session['usuario'])
     idpermiso = Permiso.objects.get(pk=int(pk))
     autorizador = Usuario.objects.filter(estado=3)

     if len(permiso.resolucion_set.all()) > 0:
        resolucion = permiso.resolucion_set.all()[0]
     else:
        resolucion = "Sin revisar"
            

     if  usuarioObj.rol.id == 1:
        if len(permiso.resolucion_set.all()) > 2:
            revisiones = Resolucion.objects.filter(permiso=permiso)
        else:
            revisiones = ""            
        #return HttpResponse(resolucion)
        return render_to_response("edt/modpermiso.html",{ "autorizador" : autorizador,"revisiones" : revisiones,"permiso_formset" : permiso_formset,"permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))

     if  usuarioObj.rol.id == 2:
        return render_to_response("edt/verpermisousuario.html",{ "resolucion" : resolucion,"permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))

def modpermisoreport(request):
    if not estaLogeado(request):
            return redirect("/login")
    usuario = Usuario.objects.get(id=request.session['usuario'])
    

    if  usuario.rol.nivel_acceso == 0:
        modificados = Bitacora.objects.filter(actividad=8)
        data = {
            "modificados":modificados,
            "usuario" : usuario,
            }
    return render_to_response("edt/modificadosreport.html",data,context_instance=RequestContext(request))


def permisos(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    permisos = Permiso.objects.all().order_by("-fecha_creacion")
    bitacoras = Bitacora.objects.all()
    revisores = Revisor.objects.all()

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                permisos = permisos.filter(usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/permisos")       
        
        paginator = Paginator(permisos,30)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            permisos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            permisos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,                 
                 "permisos": permisos,
                 "usuario" : usuarioObj,
                 "permisos_list" : permisos.object_list,
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "bitacoras" : bitacoras,
                 "revisores" : revisores,
                 }


        return render_to_response("edt/permisos.html",data,context_instance=RequestContext(request))
        #return HttpResponse(administracion)

    else:
        #if     usuarioObj.username == request.session['usuario']:
        if  usuarioObj.rol.nivel_acceso == 1:
            ids = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
            permiso = Permiso.objects.annotate(num_b=Count('resolucion')).filter(usuario=usuarioObj.id).exclude(id__in=ids).order_by("-fecha_creacion")
            

            paginator = Paginator(permiso,10)       
            try: pagina = int(request.GET.get("page",'1'))
            except ValueError: pagina = 1       
            try:
                permiso = paginator.page(pagina)
            except (InvalidPage, EmptyPage):
                permiso = paginator.page(paginator.num_pages)

        data = {
                 "foliocpe":foliocpe,
                 "folioprimaria" :folioprimaria,
                 "foliosecundaria" :foliosecundaria ,
                 "permisos": permiso,
                 "usuario" : usuario,
                 "permisos_list" : permiso.object_list,
                 "months" : mkmonth_lst(),
               }
            
        return render_to_response("edt/permisos.html",data,context_instance=RequestContext(request))

    # data = {
    #     "usuario" : usuario,
    #     "permisos": permisos,
    # }

    # return render_to_response("edt/permisos.html",data,context_instance=RequestContext(request))

def anulados(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    #anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()

    anulados = Anulado.objects.all().order_by('-fecha')
    

    bitacoras = Bitacora.objects.all()   

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                anulados = anulados.filter(permiso__usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                anulados = anulados.filter(permiso__usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/anulados")

        paginator = Paginator(anulados,10)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            anulados = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            anulados = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,
                 "usuario" : usuarioObj,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "bitacoras" : bitacoras,
                 "anulados" : anulados,
                 "query_string" : request.META["QUERY_STRING"]
                 }


        return render_to_response("edt/anuladoslst.html",data)
        
class ImprimePermisoPDF(PDFTemplateView):
    template_name="edt/imprimepermisoPDF.html"

    def get_context_data(self,**kargs):
        context = super(ImprimePermisoPDF,self).get_context_data(
            pagesize= "letter",
            title = "Impresion de Permiso",
            **kargs
            )
        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        idpermiso = self.request.GET.get("permiso")
        permiso = Permiso.objects.get(id=idpermiso)

        if len(permiso.resolucion_set.all()) > 0:
            resolucion = permiso.resolucion_set.all()[0]
        else:
            resolucion = "Sin revisar"

        context ={
                "usuario" : usuarioObj,
                "permiso" : permiso,
                "resolucion" : resolucion,
        }
        context['hoy'] =  datetime.now()

        return context

def ConPermiso(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
    rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')
    end = " "
    start = " "
    permisos = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__lte=5).exclude(id__in=anulados).exclude(id__in=rechazados).order_by("-fecha_creacion") 

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "start" in request.GET and request.GET.get("start") != "":
                permisos = permisos.filter(eventos_en_permisos__numero_evento__start__gte=request.GET.get("start"))
                start = request.GET.get("start")
                start = parser.parse(start)
                


            if "end" in request.GET and request.GET.get("end") != "":
                permisos = permisos.filter(eventos_en_permisos__numero_evento__end__lte=request.GET.get("end"))
                end = request.GET.get("end")
                end = parser.parse(end)
              

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/conpermiso")

        paginator = Paginator(permisos,30)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            permisos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            permisos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,                 
                 "permisos": permisos,
                 "usuario" : usuarioObj,
                 "end" : end,
                 "start": start, 
                 "permisos_list" : permisos.object_list,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "query_string" : request.META["QUERY_STRING"]
                 }


        return render_to_response("edt/conpermisolst.html",data)
        #return HttpResponse("hola")
    else:
        return redirect("/main")

class ConPermisoPDF(PDFTemplateView):
    template_name="edt/conpermisoPDF.html"

    def get_context_data(self,**kwargs):
        context  =super(ConPermisoPDF,self).get_context_data(
            pagesize="letter",
            title="Funcionarios con permiso",
            **kwargs
            )
        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
        rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')
        end = " "
        start = " "

        permisos = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__lte=5).exclude(id__in=anulados).exclude(id__in=rechazados).order_by("-fecha_creacion") 

        if  usuarioObj.rol.nivel_acceso == 0:
            
            estamento = Estamento.objects.all()
            usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
            estamento_filtro = Estamento.objects.all()

            if "filtrar" in self.request.GET:           

                if "start" in self.request.GET and self.request.GET.get("start") != "":
                    permisos = permisos.filter(eventos_en_permisos__numero_evento__start__gte=self.request.GET.get("start"))
                    start = self.request.GET.get("start")
                    start = parser.parse(start)
                    


                if "end" in self.request.GET and self.request.GET.get("end") != "":
                    permisos = permisos.filter(eventos_en_permisos__numero_evento__end__lte=self.request.GET.get("end"))
                    end = self.request.GET.get("end")
                    end = parser.parse(end)
                   


                if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                    estamento = self.request.GET.get("estamento")
                    permisos = permisos.filter(usuario__estamento=self.request.GET.get("estamento"))
                    for estament in estamento_filtro:
                        estament.estamento_activo = estament.id == int(estamento)

            context = {
                     "start":start,
                     "end":end,
                     "estamento":estamento,                 
                     "permisos": permisos,
                     "usuario" : usuarioObj,
                     "months" : mkmonth_lst(),
                     "usuarios_filtro" : usuarios_filtro,
                     "estamento_filtro" : estamento_filtro,
                     }
            context['hoy'] =  datetime.now()

            return context

class ConPermisoExcel(TemplateView):
    def get(self,request,*args,**kwargs):

        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
        rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')

        permisos = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__lte=5).exclude(id__in=anulados).exclude(id__in=rechazados).order_by("-fecha_creacion") 

        if  usuarioObj.rol.nivel_acceso == 0:

            
            estamento = Estamento.objects.all()
            usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
            estamento_filtro = Estamento.objects.all()

            if "filtrar" in self.request.GET:           

                if "start" in self.request.GET and self.request.GET.get("start") != "":
                    permisos = permisos.filter(eventos_en_permisos__numero_evento__start__gte=self.request.GET.get("start"))
                    start = self.request.GET.get("start")
                    start = parser.parse(start)
                    


                if "end" in self.request.GET and self.request.GET.get("end") != "":
                    permisos = permisos.filter(eventos_en_permisos__numero_evento__end__lte=self.request.GET.get("end"))
                    end = self.request.GET.get("end")
                    end = parser.parse(end)
                   


                # if "persona" in request.GET and request.GET.get("persona") != "0":
                #     persona = request.GET.get("persona")
                #     permisos = permisos.filter(usuario=request.GET.get("persona"))
                #     for usuario in usuarios_filtro:
                #         usuario.usuario_activo = usuario.id == int(persona)

                if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                    estamento = self.request.GET.get("estamento")
                    permisos = permisos.filter(usuario__estamento=self.request.GET.get("estamento"))
                    for estament in estamento_filtro:
                        estament.estamento_activo = estament.id == int(estamento)

            hoy = datetime.now()
            #Creamos el libro de trabajo
            wb= Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            #En la celda B1 ponemos el texto 'INFORME DE HORAS'
            ws['B1'] = 'FUNCIONARIOS CON PERMISO'
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('B1:E1')

            # Encabezados

            ws['B3'] = '#'
            ws['C3'] = 'Numero de Permiso'
            ws['D3'] = 'Solicitante'
            ws['E3'] = 'Fecha de Solicitud'
            ws['F3'] = 'Motivo'
            ws['G3'] = 'Inicio'
            ws['H3'] = 'Fin'
            ws['I3'] = 'Cantidad de Horas'

            cont = 4
            contador = 1

            for permiso in permisos:

                ws.cell(row=cont,column=2).value = contador
                ws.cell(row=cont,column=3).value = permiso.id
                ws.cell(row=cont,column=4).value = permiso.usuario.apellido1+" "+permiso.usuario.apellido2+" "+permiso.usuario.nombre
                ws.cell(row=cont,column=5).value = permiso.fecha_creacion
                ws.cell(row=cont,column=6).value = str(permiso.motivo)
                
                if len(permiso.eventos_en_permisos_set.all()) < 1:
                    ws.cell(row=cont,column=7).value = "Vacio"
                else: 
                    ws.cell(row=cont,column=7).value = permiso.primerEvento().numero_evento.start

                if len(permiso.eventos_en_permisos_set.all()) < 1:
                    ws.cell(row=cont,column=8).value = "Vacio"
                else:
                    ws.cell(row=cont,column=8).value = permiso.ultimoEvento().numero_evento.end

                ws.cell(row=cont,column=9).value = permiso.horas_solicitadas

                cont = cont +1
                contador = contador + 1

            nombre_archivo = "funcionarios_con_permiso.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

class AnuladosPDF(PDFTemplateView):
    template_name="edt/anuladosPDF.html"

    def get_context_data(self,**kwargs):
        context = super(AnuladosPDF,self).get_context_data(
            pagesize="letter",
            title="Anulados",
            **kwargs
            )
        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        anulados = Anulado.objects.all().order_by('-fecha')
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in self.request.GET:           

            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                anulados = anulados.filter(permiso__usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                anulados = anulados.filter(permiso__usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        context = {
                     "estamento":estamento,
                     "usuario" : usuarioObj,
                     "months" : mkmonth_lst(),
                     "usuarios_filtro" : usuarios_filtro,
                     "estamento_filtro" : estamento_filtro,                
                     "anulados" : anulados,
                     }
        context['hoy'] =  datetime.now()
        
        return context                         
        
class AnuladosExcel(TemplateView):
    """docstring for AnuladosExcel"""
    def get(self, request, *args, **kwargs):


        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        anulados = Anulado.objects.all().order_by('-fecha')
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in self.request.GET:           

            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                anulados = anulados.filter(permiso__usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                anulados = anulados.filter(permiso__usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        hoy = datetime.now()
        #Creamos el libro de trabajo
        wb= Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'INFORME DE HORAS'
        ws['B1'] = 'INFORME DE PERMISOS ANULADOS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        # Encabezados

        ws['B3'] = '#'
        ws['C3'] = 'Permiso'
        ws['D3'] = 'Solicitante'
        ws['E3'] = 'Anulador'
        ws['F3'] = 'Motivo'
        ws['G3'] = 'Fecha'

        cont = 4
        contador = 1

        for anulado in anulados:

            ws.cell(row=cont,column=2).value = contador
            ws.cell(row=cont,column=3).value = anulado.permiso.id
            ws.cell(row=cont,column=4).value = anulado.permiso.usuario.apellido1+" "+anulado.permiso.usuario.apellido2+" "+anulado.permiso.usuario.nombre
            ws.cell(row=cont,column=5).value = anulado.anuladopor.apellido1+" "+anulado.anuladopor.nombre
            ws.cell(row=cont,column=6).value = anulado.motivo
            ws.cell(row=cont,column=7).value = anulado.fecha

            cont = cont +1
            contador = contador + 1
        nombre_archivo = "informe_anulados.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response        


def descontados(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    descontados = Descontado.objects.all().order_by('-fecha')    

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                descontados = descontados.filter(permiso__usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                descontados = descontados.filter(permiso__usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/descontados")

        paginator = Paginator(descontados,10)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            descontados = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            descontados = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,
                 "usuario" : usuarioObj,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "descontados" : descontados,
                 "query_string" : request.META["QUERY_STRING"]
                 }


        return render_to_response("edt/descontadoslst.html",data)


class DescontadosPDF(PDFTemplateView):
    template_name="edt/descontadosPDF.html"

    def get_context_data(self,**kwargs):
        context = super(DescontadosPDF,self).get_context_data(
            pagesize="letter",
            title="descontados",
            **kwargs
            )
        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        descontados = Descontado.objects.all().order_by('-fecha')
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in self.request.GET:           

            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                descontados = descontados.filter(permiso__usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                descontados = descontados.filter(permiso__usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        context = {
                     "estamento":estamento,
                     "usuario" : usuarioObj,
                     "months" : mkmonth_lst(),
                     "usuarios_filtro" : usuarios_filtro,
                     "estamento_filtro" : estamento_filtro,                
                     "descontados" : descontados,
                     }
        context['hoy'] =  datetime.now()
        
        return context

class DescontadosExcel(TemplateView):
    """docstring for descontadosExcel"""
    def get(self, request, *args, **kwargs):


        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        descontados = Descontado.objects.all().order_by('-fecha')
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in self.request.GET:           

            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                descontados = descontados.filter(permiso__usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                descontados = descontados.filter(permiso__usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        hoy = datetime.now()
        #Creamos el libro de trabajo
        wb= Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'INFORME DE HORAS'
        ws['B1'] = 'INFORME DE HORAS DESCONTADAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        # Encabezados

        ws['B3'] = '#'
        ws['C3'] = 'Permiso'
        ws['D3'] = 'Solicitante'
        ws['E3'] = 'Cantidad descontada'
        ws['F3'] = 'Descontado Por'
        ws['G3'] = 'Fecha'

        cont = 4
        contador = 1

        for descontado in descontados:

            ws.cell(row=cont,column=2).value = contador
            ws.cell(row=cont,column=3).value = descontado.permiso.id
            ws.cell(row=cont,column=4).value = descontado.permiso.usuario.apellido1+" "+descontado.permiso.usuario.apellido2+" "+descontado.permiso.usuario.nombre
            ws.cell(row=cont,column=5).value = descontado.cantidad
            ws.cell(row=cont,column=6).value = descontado.ingresadopor.apellido1+" "+descontado.ingresadopor.nombre
            ws.cell(row=cont,column=7).value = descontado.fecha

            cont = cont +1
            contador = contador + 1
        nombre_archivo = "informe_descontados.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response                

def devueltos(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    devueltos = Devuelto.objects.all().order_by('-fecha')    

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                devueltos = devueltos.filter(permiso__usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                devueltos = devueltos.filter(permiso__usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/devueltos")

        paginator = Paginator(devueltos,10)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            devueltos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            devueltos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,
                 "usuario" : usuarioObj,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "devueltos" : devueltos,
                 "query_string" : request.META["QUERY_STRING"]
                 }


        return render_to_response("edt/devueltoslst.html",data)


class DevueltosPDF(PDFTemplateView):
    template_name="edt/devueltosPDF.html"

    def get_context_data(self,**kwargs):
        context = super(DevueltosPDF,self).get_context_data(
            pagesize="letter",
            title="devueltos",
            **kwargs
            )
        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        devueltos = Devuelto.objects.all().order_by('-fecha')
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in self.request.GET:           

            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                devueltos = devueltos.filter(permiso__usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                devueltos = devueltos.filter(permiso__usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        context = {
                     "estamento":estamento,
                     "usuario" : usuarioObj,
                     "months" : mkmonth_lst(),
                     "usuarios_filtro" : usuarios_filtro,
                     "estamento_filtro" : estamento_filtro,                
                     "devueltos" : devueltos,
                     }
        context['hoy'] =  datetime.now()
        
        return context

class DevueltosExcel(TemplateView):
    """docstring for devueltosExcel"""
    def get(self, request, *args, **kwargs):


        usuarioObj = Usuario.objects.get(id=self.request.session['usuario'])
        devueltos = Devuelto.objects.all().order_by('-fecha')
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in self.request.GET:           

            if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                persona = self.request.GET.get("persona")
                devueltos = devueltos.filter(permiso__usuario=self.request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                estamento = self.request.GET.get("estamento")
                devueltos = devueltos.filter(permiso__usuario__estamento=self.request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        hoy = datetime.now()
        #Creamos el libro de trabajo
        wb= Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'INFORME DE HORAS'
        ws['B1'] = 'INFORME DE HORAS DEVUELTAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        # Encabezados

        ws['B3'] = '#'
        ws['C3'] = 'Permiso'
        ws['D3'] = 'Solicitante'
        ws['E3'] = 'Cantidad devuelta'
        ws['F3'] = 'Ingresado Por'
        ws['G3'] = 'Fecha'

        cont = 4
        contador = 1

        for devuelto in devueltos:

            ws.cell(row=cont,column=2).value = contador
            ws.cell(row=cont,column=3).value = devuelto.permiso.id
            ws.cell(row=cont,column=4).value = devuelto.permiso.usuario.apellido1+" "+devuelto.permiso.usuario.apellido2+" "+devuelto.permiso.usuario.nombre
            ws.cell(row=cont,column=5).value = devuelto.cantidad
            ws.cell(row=cont,column=6).value = devuelto.ingresadopor.apellido1+" "+devuelto.ingresadopor.nombre
            ws.cell(row=cont,column=7).value = devuelto.fecha

            cont = cont +1
            contador = contador + 1
        nombre_archivo = "informe_devueltos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def upload(request, pk):
        if not estaLogeado(request):
            return redirect("/login")
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])

        BASE_DIR = os.path.dirname(os.path.dirname(__file__))
        iddocument = Document.objects.get(pk=int(pk))
        nombre = Document.objects.get(id=pk)
        
       	nombre_archivo = nombre.docfile.name
        folder = os.path.join(BASE_DIR, 'media/')
        contador = 0
        g = open(folder + nombre_archivo,'rb')
        gcal = Calendar.from_ical(g.read())
       
       
        for component in gcal.walk():         
            if component.name == "VEVENT":
                
                user = Document.objects.get(id=int(pk)).usuario.id
                nombre_funcionario = Document.objects.get(id=int(pk)).usuario.nombre
                apellido_funcionario = Document.objects.get(id=int(pk)).usuario.apellido1
                contador = contador + 1               
                CATEGORIES = component.get('CATEGORIES')
                #DTSTAMP = component.get('DTSTAMP').dt hora de creacion del archivo ics , no es relevante
                DTSTART = component.get('DTSTART').dt
                DTEND = component.get('DTEND').dt
                SUMMARY = component.get('SUMMARY')
                LOCATION = component.get('LOCATION')
                DESCRIPTION = component.get('DESCRIPTION')
                #print user," ",nombre_funcionario," ",apellido_funcionario," ",DTSTART ," ",DTEND

                c = Evento(usuario_id=user,start=DTSTART,end=DTEND)               
                c.save()  
               
        g.close()
        return render_to_response("edt/upload.html",{ "apellido_funcionario": apellido_funcionario,"nombre_funcionario":nombre_funcionario,"user" : user , "DESCRIPTION" : DESCRIPTION, "LOCATION" : LOCATION, "contador" : contador, "DTEND" : DTEND,"DTSTART": DTSTART, "CATEGORIES" : CATEGORIES,  "SUMMARY" : SUMMARY, "usuario" : usuarioObj,"nombre" : nombre})

def main(request): #antes de procesa_solicitud_permiso
    if not estaLogeado(request):
            return redirect("/login")
    hoy = datetime.now()
    this_year = hoy.year
    primaria = Estamento.objects.get(id=3)
    maternelle = Estamento.objects.get(id=2)


    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    permisos = Permiso.objects.filter(usuario=usuarioObj.id)
    bitacoras = Bitacora.objects.all()
    horass = Horas.objects.filter(usuario=usuarioObj.id).filter(permiso__fecha_creacion__year=this_year)
    formset = PermisoFormSet()

    diferencia_fechas = date.today() - usuarioObj.fecha_nac
    diferencia_fechas = diferencia_fechas.days
    edad = diferencia_fechas / 365.2425
    edad = int(edad)

    sexo = usuarioObj.sexo.id

    formset.fields["reemplazante"].queryset = Usuario.objects.filter(estado=1)

    en_sindicato = [sindi.usuario.pk for sindi in Sindicato.objects.all()]

    #pares = [ numero for numero in range(1,10) if numero % 2 == 0]
    #lista = [ {"nombre" : obj.nombre, "apellido" : obj.apelldio} for obj in queryset]   

    if usuarioObj.id not in en_sindicato:
        formset.fields["tipo"].queryset = Tipo_Permiso.objects.exclude(id__in=[3])

    usuarios = {}
    for permiso in permisos:
            idUsuario = permiso.usuario.id
            if idUsuario not in usuarios:
                usuarios[idUsuario] = {
                     "id" : permiso.usuario.id,
                     "estado" : permiso.estado,
                     "nombre" : permiso.usuario.nombre,
                     "apellido1"  :permiso.usuario.apellido1,
                     "apellido2" : permiso.usuario.apellido2,
                     "estamento" : permiso.usuario.estamento.nombre,
                     "total_horas" : 0,
                     "aprobadas" : 0,
                     "rechazadas" : 0,
                     "devolveracumuladas" : 0,
                     "devueltas" : 0,
                     "saldodevolucion" : 0,
                     "descontaracumuladas" : 0,
                     "descontadas" : 0,
                     "saldodescontar" : 0,
                     "sin_recup_con_sueldo" : 0,
                     "pendientes_por_aprobar" : 0,
                 }
             
            horas = permiso.horas_set.all()
            #traspaso de datos a otra columna
            # for hora in horas :
            #     hora.horas_por_devolver_acumuladas = hora.horas_por_devolver
            #     hora.horas_descontar_acumuladas = hora.horas_descontar
            #     hora.save()
 
            if len(horas) == 0:
                 continue
 
            hora = horas[0]
            usuarios[idUsuario]["total_horas"] += hora.horas_solicitadas
            usuarios[idUsuario]["aprobadas"] += hora.horas_aprobadas
            usuarios[idUsuario]["rechazadas"] += hora.horas_rechazadas
            usuarios[idUsuario]["devolveracumuladas"] += hora.horas_por_devolver_acumuladas
            usuarios[idUsuario]["devueltas"] += hora.horas_devueltas
            usuarios[idUsuario]["saldodevolucion"] += hora.horas_por_devolver
            usuarios[idUsuario]["descontaracumuladas"] += hora.horas_descontar_acumuladas
            usuarios[idUsuario]["descontadas"] += hora.horas_descontadas
            usuarios[idUsuario]["saldodescontar"] += hora.horas_descontar
            usuarios[idUsuario]["sin_recup_con_sueldo"] += hora.horas_sin_recuperacion_con_goce
            usuarios[idUsuario]["pendientes_por_aprobar"] += hora.horas_pendientes_por_aprobar
         
    usuariosLista = [value for key,value in usuarios.iteritems()]
   
    data = {
            "sexo" : sexo,
            "edad" : edad,
            "form": formset,
            "usuario": usuarioObj,
            "permisos" : permisos ,
            "bitacoras" : bitacoras,
            "usuarios" : usuariosLista,
            "horas" : horass,
            "en_sindicato" : en_sindicato,
            "hoy": hoy,
            "primaria" : primaria,
            "maternelle" : maternelle,
            #"fech" : fech,
           }

    
    return render_to_response("edt/main.html", data)
   # return render_to_response("edt/main.html", {"user" : usuarioObj},context_instance=RequestContext(request))

def formacion(request):
    if not estaLogeado(request):
            return redirect("/login")
    hoy = datetime.now()
    agno = hoy.year
    flag = 0
    usuario = Usuario.objects.get(id=request.session['usuario'])
    data = {
        "usuario": usuario,
        }

    diferencia_fechas = date.today() - usuario.fecha_nac
    diferencia_fechas = diferencia_fechas.days
    edad = diferencia_fechas / 365.2425
    edad = int(edad)


    if request.method == 'POST':
        nombre = request.POST.get("nombre")
        inicio = request.POST.get("inicio")
        fin = request.POST.get("fin")
        ubicacion = request.POST.get("ubicacion")
        flag = 1

        formacion = Formacion(nombre=nombre,inicio=inicio,fin=fin,ubicacion=ubicacion,agno=agno)               
        formacion.save()  

        

        data = {
            "edad" : edad,
            "usuario": usuario,
            "nombre" : nombre,
            "inicio" : inicio,
            "fin" :  fin,
            "ubicacion" : ubicacion,
            "flag" : flag,
            "agno" :  agno,
        }

        #formacion = Formacion(nombre=nombre,inicio=inicio,fin=fin,ubicacion=ubicacion,agno=agno)
        #formacion.save()

        #return HttpResponse(edad)
        return render_to_response("edt/formacion.html",data,context_instance=RequestContext(request))
    else:
        return render_to_response("edt/formacion.html",data,context_instance=RequestContext(request))    

def formacionlist(request):
    if not estaLogeado(request):
        return redirect("/login")

    usuario = Usuario.objects.get(id=request.session['usuario'])
    formaciones = Formacion.objects.all()

    data = {
            "usuario" :  usuario,
            "formaciones" : formaciones, 
           }
    return render_to_response("edt/formacionlist.html",data,context_instance=RequestContext(request))

def formacionEdit(request,pk):
    if not estaLogeado(request):
            return redirect("/login")
    hoy = datetime.now()
    agno = hoy.year
    flag = 0
    usuario = Usuario.objects.get(id=request.session['usuario'])    
    formacion = Formacion.objects.get(id=pk)

    if request.POST:
        formacion.nombre = request.POST.get('nombre')
        formacion.ubicacion = request.POST.get('ubicacion')
        formacion.inicio = request.POST.get('inicio')
        formacion.fin = request.POST.get('fin')
        formacion.save()
        flag = 1

    data = {

        "formacion" : formacion,
        "usuario" :  usuario,
        "flag" : flag,
    }
         

    return render_to_response("edt/formacionEdit.html",data,context_instance=RequestContext(request))


def salidapedagogica(request):    
    if not estaLogeado(request):
            return redirect("/login")
    hoy = datetime.now()
    agno = hoy.year
    flag = 0
    usuario = Usuario.objects.get(id=request.session['usuario'])
    data = {
        "usuario": usuario,
    }
    

    diferencia_fechas = date.today() - usuario.fecha_nac
    diferencia_fechas = diferencia_fechas.days
    edad = diferencia_fechas / 365.2425
    edad = int(edad)


    if request.method == 'POST':
        nombre = request.POST.get("nombre")
        inicio = request.POST.get("inicio")
        fin = request.POST.get("fin")
        ubicacion = request.POST.get("ubicacion")
        flag = 1

        salidapedagogica = SalidaPedagogica(nombre=nombre,inicio=inicio,fin=fin,ubicacion=ubicacion,agno=agno)
        salidapedagogica.save()

        data = { 
            "edad" : edad,
            "usuario": usuario,
            "nombre" : nombre,
            "inicio" : inicio,
            "fin" :  fin,
            "ubicacion" : ubicacion,
            "flag" : flag,
            "agno" :  agno,
        }

        #return HttpResponse(inicio)
        return render_to_response("edt/SalidaPedagogica.html",data,context_instance=RequestContext(request))
    else:
        return render_to_response("edt/SalidaPedagogica.html",data,context_instance=RequestContext(request))

def SalidaPedagogicalist(request):
    if not estaLogeado(request):
        return redirect("/login")

    usuario = Usuario.objects.get(id=request.session['usuario'])
    SalidaPedagogicas = SalidaPedagogica.objects.all()

    data = {
            "usuario" :  usuario,
            "SalidaPedagogicas" : SalidaPedagogicas, 
           }
    return render_to_response("edt/SalidaPedagogicalist.html",data,context_instance=RequestContext(request))

def SalidaPedagogicaEdit(request,pk):
    if not estaLogeado(request):
            return redirect("/login")
    hoy = datetime.now()
    agno = hoy.year
    flag = 0
    usuario = Usuario.objects.get(id=request.session['usuario'])    
    Salida = SalidaPedagogica.objects.get(id=pk)

    if request.POST:
        Salida.nombre = request.POST.get('nombre')
        Salida.ubicacion = request.POST.get('ubicacion')
        Salida.inicio = request.POST.get('inicio')
        Salida.fin = request.POST.get('fin')
        Salida.save()
        flag = 1

    data = {

        "Salida" : Salida,
        "usuario" :  usuario,
        "flag" : flag, 
    }

    return render_to_response("edt/SalidaPedagogicaEdit.html",data,context_instance=RequestContext(request))


def AusenciaLaboral(request):
    if not estaLogeado(request):
            return redirect("/login")
    hoy = datetime.now()
    agno = hoy.year
    flag = 0
    opt = 0
    hoy = datetime.now()
    agno = hoy.year        
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    funcionarios = Usuario.objects.all().filter(estado=1)
    reemplazantes = Usuario.objects.all().filter(estado=1)
    formaciones = Formacion.objects.all().filter(inicio__year=agno)
    salidaspedagogicas = SalidaPedagogica.objects.all().filter(inicio__year=agno)
    motivos = Motivo_Ausencia_Laboral.objects.all()

    if request.POST:
        flag = 1
        opt = request.POST.get('motivo')
        funcionario = request.POST.get("funcionario")
        reemplazante = request.POST.get("reemplazante")
        funcionario_ = Usuario.objects.get(id=funcionario)
        reemplazante_ = Usuario.objects.get(id=reemplazante)
        motivo = Motivo_Ausencia_Laboral.objects.get(id=opt)

        if opt == '1' :            
            formacion = request.POST.get("formacion")              
            formacion_ = Formacion.objects.get(id=formacion)
            ausencia = Ausencia_Laboral(funcionario=funcionario_, formacion=formacion_, reemplazante=reemplazante_, motivo=motivo)
            ausencia.save()

        if opt == '2':             
            salida = request.POST.get("salida")
            salida_ = SalidaPedagogica.objects.get(id=salida)                     
            ausencia = Ausencia_Laboral(funcionario=funcionario_, salida=salida_, reemplazante=reemplazante_, motivo=motivo)
            ausencia.save()


        if opt >= '3':
            comentario = request.POST.get("comentario")
            inicio = request.POST.get("inicio")
            fin = request.POST.get("fin")
            ausencia = Ausencia_Laboral(funcionario=funcionario_, comentario=comentario, inicio=inicio, fin=fin, reemplazante=reemplazante_, motivo=motivo)
            ausencia.save()


    
    data = {
        "usuario" : usuarioObj,
        "funcionarios" : funcionarios,
        "reemplazantes" :  reemplazantes,
        "formaciones" : formaciones,
        "salidaspedagogicas" : salidaspedagogicas,
        "motivos" : motivos,
        "flag" : flag,
        "opt" : opt,
        "agno" : agno,
    }


    return render_to_response("edt/ausencia_laboral.html",data,context_instance=RequestContext(request))








def descontar(request):
    if not estaLogeado(request):
            return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])    
    revisados = Resolucion.objects.values_list('permiso')        
    anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
    rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')
    condevolucion= Permiso.objects.filter(devuelve_horas='S')
   

    if  usuarioObj.rol.nivel_acceso == 0: 
        usuarios = Usuario.objects.all().filter(estado=1) #.exclude(permiso__horas__horas_solicitadas=None)

        if request.method == 'POST':
            if "usuario" in request.POST and request.POST.get("usuario") != "":
                seleccionado = request.POST.get("usuario")
                permisos  = Permiso.objects.filter(usuario=seleccionado)
                horas = Horas.objects.filter(permiso__usuario=seleccionado)
                horas = horas.exclude(permiso__id__in=anulados)
                horas = horas.filter(permiso__id__in=revisados)
                horas = horas.exclude(permiso__id__in=rechazados)
                horas = horas.exclude(permiso__id__in=condevolucion)
               
                for user in usuarios:
                    user.usuario_activo = user.id == int(seleccionado)

            if "horasdescontar" in request.POST and request.POST.get("horasdescontar") != "":
                permisoid = request.POST.get("permisoid")
                horasdescontar = request.POST.get("horasdescontar")
                permisodesc  = Permiso.objects.get(id=permisoid)              
                descontado = Descontado(permiso=permisodesc,cantidad=horasdescontar,ingresadopor=usuarioObj)
                descontado.save()
                if len(permisodesc.horas_set.all()) > 0:
                    horas = Horas.objects.get(permiso=permisoid)
                    horas.horas_descontadas = float(horasdescontar)
                    horas.horas_descontar = float(horas.horas_descontar) - float(horasdescontar) 
                    horas.save()
                    data = {
                        "usuario": usuarioObj,
                        "horasdescontar" : horasdescontar,
                        "permiso" : permisodesc,

                    }

                    return render_to_response("edt/horasdescontadas.html",data,context_instance=RequestContext(request))
                    #return HttpResponse(horasdescontar)
                else:
                    return HttpResponse("mayor que 0")    

                return HttpResponse(horasdescontar)    

            return render_to_response("edt/descontarhoras.html",{"horas" : horas,"usuarios" : usuarios,"permisos":permisos,"usuario"  :usuarioObj },context_instance=RequestContext(request))    

            #return HttpResponse(permisos)
            #return redirect('/devueltas/%d'%(horas.id))
        else:                
                
            return render_to_response("edt/descontarhoras.html",{"anulados" : anulados,"usuarios": usuarios,"usuario": usuarioObj},context_instance=RequestContext(request))
    
    else:
        return redirect("/main")

def devuelvehoras(request):
    if not estaLogeado(request):
            return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    revisados = Resolucion.objects.values_list('permiso')        
    anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
    rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')
    sindevolucion= Permiso.objects.filter(devuelve_horas='N')
   

    if  usuarioObj.rol.nivel_acceso == 0: 
        usuarios = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)

        if request.method == 'POST':
            if "usuario" in request.POST and request.POST.get("usuario") != "":
                seleccionado = request.POST.get("usuario")
                permisos  = Permiso.objects.filter(usuario=seleccionado)
                horas = Horas.objects.filter(permiso__usuario=seleccionado)
                horas = horas.exclude(permiso__id__in=anulados)
                horas = horas.filter(permiso__id__in=revisados)
                horas = horas.exclude(permiso__id__in=rechazados)
                horas = horas.exclude(permiso__id__in=sindevolucion)

                for user in usuarios:
                    user.usuario_activo = user.id == int(seleccionado)

            if "horasdevolver" in request.POST and request.POST.get("horasdevolver") != "":
                permisoid = request.POST.get("permisoid")
                horasdevueltas = request.POST.get("horasdevolver")
                permisodesc  = Permiso.objects.get(id=permisoid)                
                devuelto = Devuelto(permiso=permisodesc,cantidad=horasdevueltas,ingresadopor=usuarioObj)
                devuelto.save()

                if len(permisodesc.horas_set.all()) > 0:                    
                    horas = Horas.objects.get(permiso=permisoid)

                    horas.horas_por_devolver = float(horas.horas_por_devolver) - float(horasdevueltas)

                    horas.horas_devueltas = float(horasdevueltas) + float(horas.horas_devueltas)

                    horas.save()
                    data = {
                        "usuario": usuarioObj,
                        "horasdevueltas" : horasdevueltas,
                        "permiso" : permisodesc,                        
                    }

                    return render_to_response("edt/devueltas.html",data,context_instance=RequestContext(request))
                    #return HttpResponse(horasdescontar)
                else:
                    return HttpResponse("mayor que 0")    

                return HttpResponse(horasdevolver)    

            return render_to_response("edt/devuelvehoras.html",{"horas" : horas,"usuarios" : usuarios,"permisos":permisos,"usuario"  :usuarioObj },context_instance=RequestContext(request))    

            #return HttpResponse(permisos)
            #return redirect('/devueltas/%d'%(horas.id))
        else:                
                
            #return HttpResponse(anulados)
            return render_to_response("edt/devuelvehoras.html",{"anulados" : anulados,"usuarios": usuarios,"usuario": usuarioObj},context_instance=RequestContext(request))
    
    else:
        return redirect("/main")
   
def devueltas(request,pk):
    if not estaLogeado(request):
            return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    horas = Horas.objects.get(pk=int(pk))

    return render_to_response("edt/devueltas.html",{"horas": horas,"usuario": usuarioObj},context_instance=RequestContext(request))

def mkmonth_lst():

    if not Permiso.objects.count(): return[]

    year, month = time.localtime()[:2]
    first = Permiso.objects.order_by("fecha_creacion")[0]
    fyear = first.fecha_creacion.year
    fmonth = first.fecha_creacion.month
    months = []

    for y in range(year,fyear-1,-1):
        start,end = 12,0
        if y == year:start = month
        if y == fyear: end = fmonth -1

        for m in range(start,end,-1):
            months.append((y,m,month_name[m]))
        
    return months

def month(request,year,month):
    entrada = Permiso.objects.filter(fecha__year=year,fecha__month=month)
    return render_to_response("edt/permisolst.html",dict(permiso_list=permiso,user=request.user,month=mkmonth_lst(),archive=True))

def verpermiso(request, pk):
     if not estaLogeado(request):
            return redirect("/login")
     else:
         permiso = Permiso.objects.get(id=pk)
         permiso_formset = PermisoFormSet(instance=permiso)
         usuarioObj = Usuario.objects.get(id=request.session['usuario'])
         idpermiso = Permiso.objects.get(pk=int(pk))         
         if len(permiso.resolucion_set.all()) > 0:
            resolucion = permiso.resolucion_set.all()[0]
         else:
            resolucion = "Sin revisar"
            

     if  usuarioObj.rol.id == 1:
        bitacora = Bitacora.objects.annotate(contador=Count('actividad')).filter(permiso=permiso).filter(actividad=2)
        cuenta = len(bitacora)

        if len(permiso.resolucion_set.all()) > 0:
            revisiones = Resolucion.objects.filter(permiso=permiso)
        else:
            revisiones = ""            
        #return HttpResponse(resolucion)
        return render_to_response("edt/verpermiso.html",{"cuenta" : cuenta,"bitacora" : bitacora, "revisiones" : revisiones,"permiso_formset" : permiso_formset,"permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))

     if  usuarioObj.rol.id == 2:
        return render_to_response("edt/verpermisousuario.html",{ "resolucion" : resolucion,"permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))

def verpermiso2(request, pk):
     if not estaLogeado(request):
            return redirect("/login")
     else:
         permiso = Permiso.objects.get(id=pk)
         permiso_formset = PermisoFormSet(instance=permiso)
         usuarioObj = Usuario.objects.get(id=request.session['usuario'])
         idpermiso = Permiso.objects.get(pk=int(pk))         
         if len(permiso.resolucion_set.all()) > 0:
            resolucion = permiso.resolucion_set.all()[0]
         else:
            resolucion = "Sin revisar"
            

     if  usuarioObj.rol.id == 1:
        if len(permiso.resolucion_set.all()) > 0:
            revisiones = Resolucion.objects.filter(permiso=permiso)
        else:
            revisiones = ""            
        #return HttpResponse(resolucion)
        return render_to_response("edt/verpermiso2.html",{ "revisiones" : revisiones,"permiso_formset" : permiso_formset,"permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))

     if  usuarioObj.rol.id == 2:
        return render_to_response("edt/verpermisousuario.html",{ "resolucion" : resolucion,"permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))



def ingresapermiso(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])     

    if request.method == 'POST':
        formset = PermisoFormSet(request.POST, request.FILES)           
        if formset.is_valid():
            permiso= formset.save(commit=False)
            permiso.usuario = usuarioObj
            permiso.save()
            return redirect("/comprobante/%d"%(permiso.id))

            # do something
    else:
        formset = PermisoFormSet()
    return render_to_response("edt/permiso.html", {"form": formset,})

    
def mostrar_respuesta(request, pk):
    if not estaLogeado(request):
        return redirect("/login")
    else:   
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])
        idresolucion = Resolucion.objects.get(pk=int(pk))
        resolucion = Resolucion.objects.get(id=pk)
        permiso = resolucion.permiso.id

        data ={

            "idresolucion" : idresolucion,
            "usuario" : usuarioObj,
            "permiso" : permiso,
            "resolucion" : resolucion,
        }
    
    return render_to_response("edt/resolucion.html",data,context_instance=RequestContext(request)) 

def aprobarRechazar(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])     
    

    if request.POST:
        permiso = Permiso.objects.get(id=request.POST['permiso'])
        bitacora = Bitacora.objects.annotate(contador=Count('actividad')).filter(permiso=permiso).filter(actividad=2)
        aprovaciones = len(bitacora)        

        if request.POST['respuesta'] == 'A' :
            estado_permiso = Estado_Permiso.objects.get(id=4)            
            permiso.comentario = request.POST.get('comentario')
            permiso.sueldo = request.POST.get('sueldo')
            permiso.devuelve_horas = request.POST.get('devuelve_horas')
            permiso.estado = estado_permiso
            #permiso.estado = Estado_Permiso.objects.filter(id=4) 
            permiso.save()
            horas = Horas.objects.get(permiso=permiso)            
            #horas = Horas(permiso=permiso,usuario=permiso.usuario,horas_solicitadas=permiso.horas_solicitadas,horas_aprobadas=permiso.horas_solicitadas,horas_por_devolver=permiso.horas_solicitadas)
            horas.permiso = permiso
            horas.usuario = permiso.usuario
            horas.horas_solicitadas = permiso.horas_solicitadas
            horas.horas_aprobadas = permiso.horas_solicitadas
            horas.horas_pendientes_por_aprobar = 0
            
            if aprovaciones == 0:

                if permiso.devuelve_horas == 'S':
                    horas.horas_por_devolver = permiso.horas_solicitadas
                    horas.horas_por_devolver_acumuladas = permiso.horas_solicitadas
                
                if (permiso.devuelve_horas == 'N' and permiso.sueldo == 'S'):          
                    horas.horas_descontar = permiso.horas_solicitadas
                    horas.horas_por_devolver = 0                 

                if (permiso.devuelve_horas == 'N' and permiso.sueldo == 'C'):
                    horas.horas_descontar = 0
                    horas.horas_sin_recuperacion_con_goce = permiso.horas_solicitadas


            horas.save()

            form = PermisoFormSetEdit(request.POST,instance=permiso)
            if form.is_valid():
                form.save()
            actividad = Actividad.objects.get(id=2)
            resu = procesa_resolucion(request,actividad,usuarioObj)
            

        if request.POST['respuesta'] == 'R' :
            
            horas = Horas.objects.get(permiso=permiso)
            estado_permiso = Estado_Permiso.objects.get(id=5)
            permiso.estado = estado_permiso
            permiso.save()
            #horas = Horas(permiso=permiso,usuario=permiso.usuario,horas_solicitadas=permiso.horas_solicitadas,horas_rechazadas=permiso.horas_solicitadas)
            horas.permiso = permiso
            horas.usuario = permiso.usuario
            horas.horas_horas_solicitadas = permiso.horas_solicitadas
            horas.horas_rechazadas = permiso.horas_solicitadas
            horas.horas_pendientes_por_aprobar = 0
            horas.save()
            actividad = Actividad.objects.get(id=3)
            resu = procesa_resolucion(request,actividad,usuarioObj)        


        if request.POST['respuesta'] == 'M' :
            estado_permiso = Estado_Permiso.objects.get(id=4)            
            permiso.comentario = request.POST.get('comentario')
            permiso.sueldo = request.POST.get('sueldo')
            permiso.devuelve_horas = request.POST.get('devuelve_horas')
            permiso.estado = estado_permiso
            #permiso.estado = Estado_Permiso.objects.filter(id=4) 
            permiso.save()
            horas = Horas.objects.get(permiso=permiso)            
            #horas = Horas(permiso=permiso,usuario=permiso.usuario,horas_solicitadas=permiso.horas_solicitadas,horas_aprobadas=permiso.horas_solicitadas,horas_por_devolver=permiso.horas_solicitadas)
            horas.permiso = permiso
            horas.usuario = permiso.usuario
            horas.horas_solicitadas = permiso.horas_solicitadas
            horas.horas_aprobadas = permiso.horas_solicitadas
            horas.horas_pendientes_por_aprobar = 0
            if permiso.devuelve_horas == 'S':
                horas.horas_por_devolver = permiso.horas_solicitadas
                horas.horas_por_devolver_acumuladas = permiso.horas_solicitadas                

            if (permiso.devuelve_horas == 'N' and permiso.sueldo == 'S'):          
                horas.horas_descontar = permiso.horas_solicitadas
                horas.horas_por_devolver = 0                 

            if (permiso.devuelve_horas == 'N' and permiso.sueldo == 'C'):
                horas.horas_descontar = 0
                horas.horas_sin_recuperacion_con_goce = permiso.horas_solicitadas


            horas.save()

            form = PermisoFormSetEdit(request.POST,instance=permiso)
            if form.is_valid():
                form.save()
            actividad = Actividad.objects.get(id=8)
            resu = procesa_modificacion(request,actividad,usuarioObj)


        return redirect("/respuesta/%d"%(resu.id))



def anularlst(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    anulados = Bitacora.objects.values_list("permiso").filter(actividad__id=4).distinct()
    rechazados = Resolucion.objects.values_list("permiso").filter(respuesta='R')

    

    permisos = Permiso.objects.annotate(num_b=Count('resolucion')).filter(num_b__lte=5).exclude(id__in=anulados).exclude(id__in=rechazados).order_by("-fecha_creacion") 

    cpe = Usuario.objects.values_list("jefatura").filter(jefatura__id=1).filter(estado=1)
    foliocpe = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=cpe))
    
    primaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=4).filter(estado=1)
    folioprimaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=primaria))

    secundaria = Usuario.objects.values_list("jefatura").filter(jefatura__id=5).filter(estado=1)
    foliosecundaria = len(Permiso.objects.annotate(sec=Count('usuario')).filter(usuario__jefatura=secundaria))

    if  usuarioObj.rol.nivel_acceso == 0:

        
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().exclude(permiso__horas__horas_solicitadas=None).filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                permisos = permisos.filter(usuario=request.GET.get("persona"))
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                permisos = permisos.filter(usuario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)   
        elif "limpiar" in request.GET:
            return redirect("/anularlst")  

        paginator = Paginator(permisos,30)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            permisos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            permisos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,
                 "foliocpe":foliocpe,
                 "folioprimaria" :folioprimaria,
                 "foliosecundaria" :foliosecundaria ,
                 "permisos": permisos,
                 "usuario" : usuarioObj,
                 "permisos_list" : permisos.object_list,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 }


        return render_to_response("edt/anularlst.html",data)
    else:
        return redirect("/main")

def anulapermiso(request,pk):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    idpermiso = Permiso.objects.get(pk=int(pk))
    permiso = Permiso.objects.get(id=pk)

    return render_to_response("edt/anulapermiso.html",{ "permiso" : idpermiso,"usuario" : usuarioObj},context_instance=RequestContext(request))

def anula(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    estado_permiso = Estado_Permiso.objects.get(id=6)

    if request.POST:        

        if request.POST.get('anular') == 'Anulado' :

            permiso_id = request.POST.get('permiso')
            motivo = request.POST.get('motivo')
            #guardo en permiso reseteando las horas a 0 
            permiso = Permiso.objects.get(id=permiso_id)
            # permiso.horas_solicitadas = 0
            # permiso.horas_solicitadas_funcionario = 0
            permiso.estado = estado_permiso
            permiso.save()
            #guardo en Bitacora
            actividad = Actividad.objects.get(id=4)
            bitacora = Bitacora(actividad=actividad,usuario=usuarioObj,permiso=permiso)            
            bitacora.save()
            #guardo en Anulado para el registro de los permisos anulados
            anulado = Anulado(permiso=permiso,anuladopor=usuarioObj,motivo=motivo)
            anulado.save()
            #guardo en Resolucion
            resolucion = Resolucion(respuesta='N',razon=motivo,permiso=permiso,resolutor=usuarioObj)
            resolucion.save()

            #arrayAnular = []

            para_anular = Eventos_en_Permisos.objects.filter(numero_permiso=permiso) 

            for evento_anulado in para_anular :#aqui
                anulados = Eventos_en_Permisos_Anulados(evento=evento_anulado.numero_evento,permiso=permiso,deltafuncionario=evento_anulado.deltafuncionario,deltainforme=evento_anulado.deltainforme)
                anulados.save()
                
            
            #libero los eventos para poder ser reutilizados
            eventos = Eventos_en_Permisos.objects.filter(numero_permiso=permiso).delete()            
            #guardo en horas reseteando las horas_solicitadas a 0 
            horas = Horas.objects.get(permiso=permiso)
            horas.horas_solicitadas = 0
            horas.horas_descontar = 0
            horas.horas_por_devolver = 0
            horas.horas_aprobadas = 0
            horas.horas_pendientes_por_aprobar = 0 
            horas.save()


        if request.POST.get('cancelar') == 'Cancelado':

            return redirect("/anularlst")    
            
    return redirect("/anulado/%d"%(bitacora.id))


def mostrar_anulado(request,pk):
    if not estaLogeado(request):
        return redirect("/login")

    else:   
        usuarioObj = Usuario.objects.get(id=request.session['usuario'])
        idbitacora = Bitacora.objects.get(pk=int(pk))
        bitacora = Bitacora.objects.get(id=pk)
        permiso = bitacora.permiso.id    

    
    return render_to_response("edt/anulado.html",{ "bitacora" : bitacora, "usuario" : usuarioObj,"permiso" : permiso},context_instance=RequestContext(request)) 

def horas(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario']) 

    if  usuarioObj.rol.nivel_acceso == 0:

        #Calculo de Horas Solicitadas, Devueltas por Usuario
        horas =  Usuario.objects.annotate(horas_sol=Sum("horas__horas_solicitadas")).annotate(horas_dev=Sum("horas__horas_devueltas")).annotate(total=F('horas_sol') - F('horas_dev')).annotate(h=Count('horas')).exclude(h=0).order_by("apellido1").filter(estado=1)

        paginator = Paginator(horas,10)       
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1       
        try:
            permiso = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            permiso = paginator.page(paginator.num_pages)

        return render_to_response("edt/horas.html",{"usuario": usuarioObj,"horas":horas,"permisoObj_list" : permiso.object_list,"months" : mkmonth_lst()},context_instance=RequestContext(request))
    else:
        return redirect("/main")


def reemplazolicencia(request):
    if not estaLogeado(request):
                return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])
    hoy = datetime.now()
    agno = hoy.year
    licencias = Licencia.objects.all().values('funcionario').order_by("-inicio")
    
    if  user.rol.id == 1:
        usuarios_filtro = Usuario.objects.all().filter(estado=1).filter(id__in=licencias)
        reemplazantes_filtro = Usuario.objects.all().filter(estado=1)
        paga = " "
        horasreemplazo = 0
        horaspagar = 0
        fecha = 0
        lic = " "
        licen = " "
        funcionario = ""
        licencias_filtro = ""
        flag = 0

        if "funcionario" and not "limpiar"  in request.GET:

            flag = 1

            if "funcionario" in request.GET and request.GET.get("funcionario") != "0":
                funcionario = request.GET.get("funcionario")
                licencias_filtro = Licencia.objects.all().filter(funcionario_id=funcionario) 

                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(funcionario)


                if "reemplazante" in request.GET and request.GET.get("reemplazante") != "0":
                    reemplazante = request.GET.get("reemplazante")
                    reemplazantee = Usuario.objects.filter(id=reemplazante)
                    for r in reemplazantee:
                        reemplaza = r
                                    
                    for reemplazant in reemplazantes_filtro:
                        reemplazant.reemplazante_activo = reemplazant.id == int(reemplazante)

                if "licencia" in request.GET and request.GET.get("licencia") !="0":
                    licencia = request.GET.get("licencia")
                    
                    for licen in licencias_filtro:
                        licen.licencia_activa = licen.id == int(licencia)

                    #licenci = Licencia.objects.all().filter(id=licencia)
 
                # if "paga" in request.GET and request.GET.get("paga") !="0":
                #     paga = request.GET.get("paga")
                    

                if "horasreemplazo" in request.GET and request.GET.get("horasreemplazo") !=0:
                    horasreemplazo = request.GET.get("horasreemplazo")

                if "horaspagar" in request.GET and request.GET.get("horaspagar") >= 0:
                    horaspagar = request.GET.get("horaspagar")    
                    
                
                if "fecha" in request.GET and request.GET.get("fecha") !=" ":
                    fecha = request.GET.get("fecha")
                        

                if  "reemplazante" in request.GET and request.GET.get("reemplazante") != "0" and "licencia" in request.GET and request.GET.get("licencia") != "0" and "horasreemplazo" in request.GET and request.GET.get("horasreemplazo") !=0 and "horaspagar" in request.GET and request.GET.get("horaspagar") !=0 and "fecha" in request.GET and request.GET.get("fecha") !=" ":

                    reemplazolicencia = ReemplazoLicencia(licencia=licen,reemplazante=reemplaza,horasreemplazo=horasreemplazo, horaspagar=horaspagar,fecha=fecha, ingresadopor=user)
                    reemplazolicencia.save()

                    return redirect("/ingresoreemplazo/%d"%(reemplazolicencia.id))

        elif "limpiar" in request.GET:
            return redirect("/reemplazolicencia")
            
        data = {
            "funcionario": funcionario,
            "usuarios_filtro" : usuarios_filtro,
            "perro" : "perro",
            "usuario":user,
            "licencias_filtro" : licencias_filtro,
            #"licenci" : licenci,
            "paga" : paga,
            "horasreemplazo" : horasreemplazo,
            "horaspagar" : horaspagar,
            "fecha" : fecha,
            "reemplazantes_filtro" : reemplazantes_filtro,
            "flag" : flag,

        }

    return render_to_response("edt/reemplazolicencia.html",data,context_instance=RequestContext(request))

def ingresoreemplazo(request,pk):
    if not estaLogeado(request):
                return redirect("/login")
    
    user = Usuario.objects.get(id=request.session['usuario'])
    reemplazo = ReemplazoLicencia.objects.get(id=pk)

    data = {
            "reemplazo": reemplazo,
            "usuario" : user,
    }

    return render_to_response("edt/comprobantereemplazo.html",data,context_instance=RequestContext(request))


def validalicencia(request):
    
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    funcionario = request.GET.get("funcionario")

    registrado = Licencia.objects.filter(
        funcionario__id=funcionario,
        inicio__lte=fin,
        fin__gte=inicio
    ).exists()

    variable_ajax = {
        "registrado" : registrado
    }

    return HttpResponse(json.dumps(variable_ajax),content_type='application/json')            
            


def licencia(request):
    if not estaLogeado(request):
        return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])
    flag = ''
    inactivos = Usuario.objects.all().filter(estado=2)

    if  user.rol.id == 1:
        usuarios_filtro = Usuario.objects.all().exclude(id__in=inactivos)

        if request.method == 'POST':
            formset = LicenciaFormset(request.POST, request.FILES)
            #licenciafuncionario = Licencia.objects.filter(funcionario=request.POST.get("funcionario"))#se obtiene licencias de funcionario selecionado para validar

            if formset.is_valid():
                #if request.POST.get("inicio") == 

                licencia = formset.save(commit=False)
                licencia.inicio = request.POST.get("inicio")
                licencia.inicio = datetime.strptime(licencia.inicio + " 00:01:00"  , "%Y-%m-%d %H:%M:%S")
                licencia.fin = request.POST.get("fin")
                licencia.fin = datetime.strptime(licencia.fin  + " 00:01:00" , "%Y-%m-%d %H:%M:%S")
                licencia.ingresadopor = user
                licencia.cantidad_dias = request.POST.get("cantidad_dias")
                licencia.save()
                
                licencias = Licencia.objects.all().order_by('-id')[0] #obtencion la última licencia ingresada

                eventos = Evento.objects.filter(usuario=licencia.funcionario).filter(start__gte=licencia.inicio).filter(end__lte=licencia.fin)
                suma = 0
                sumafuncionario = 0
                i = 0
                deltas = []
                deltaf = [] 

                for evento in eventos :
                    delta = evento.end - evento.start # calculo de la cantidad de horas solicitadas en segundos
            
                    if (licencias.funcionario.estamento.id == 9 ):  
                    #CALCULO PARA SECUNDARIA
                        suma += float(delta.seconds) / 3300 # calculo para  Informes
                        deltag = float(delta.seconds) / 3300 # calculo para  Informes
                        deltas.append(round(deltag,2))
                        sumafuncionario += float(delta.seconds) / 3300  # calculo para  funcionario
                        deltafuncionario = float(delta.seconds) / 3300  # calculo para  funcionario
                        deltaf.append(round(deltafuncionario,2))  
                      
                        suma = round(suma,2) # redondeo a 2 decimales
                    


                    if (licencias.funcionario.estamento.id == 5 or licencias.funcionario.estamento.id == 4 ):  
                        #CALCULO PARA SECUNDARIA
                        suma += float(delta.seconds) / 3300 # calculo para  Informes
                        deltag = float(delta.seconds) / 3300 # calculo para  Informes
                        deltas.append(round(deltag,2))
                        sumafuncionario += float(delta.seconds) / 3300  # calculo para  funcionario
                        deltafuncionario = float(delta.seconds) / 3300  # calculo para  funcionario
                        deltaf.append(round(deltafuncionario,2))  
                          
                        suma = round(suma,2) # redondeo a 2 decimales
                        
                        

                    if (licencias.funcionario.estamento.id == 2 or licencias.funcionario.estamento.id == 3):
                        #CALCULO DE HORAS PARA PRIMARIA
                        suma += float(delta.seconds) / 2700
                        deltag = float(delta.seconds) / 2700
                        deltas.append(round(deltag,2))
                        suma = round(suma,2) # redondeo a 2 decimales
                        sumafuncionario += float(delta.seconds) / 2700   # calculo para  funcionario
                        deltafuncionario = float(delta.seconds) / 2700  # calculo para  funcionario
                        deltaf.append(round(deltafuncionario,2))
                        

                    if (licencias.funcionario.estamento.id == 1 or licencias.funcionario.estamento.id == 6 or licencias.funcionario.estamento.id == 7 or licencias.funcionario.estamento.id == 8):
                        #CALCULO DE HORAS PARA ADMINISTRACION,ASEM
                        suma += float(delta.seconds) / 3600  
                        deltag = float(delta.seconds) / 3600
                        deltas.append(round(deltag,2))
                        suma = round(suma,2) # redondeo a 2 decimales
                        sumafuncionario = suma


                licencia.horas = suma
                licencia.save()

                #============================================= envio de emails cuando se ingresa una licencia================
                #=============================================definicion de destinatarios
                proviseur = 'jbouthier@cdegaulle.cl'
                dirgen = 'dirgen@cdegaulle.cl'
                secondaire = 'secondaire@cdegaulle.cl'
                director_primaria = 'vvanrell@cdegaulle.cl'
                director_secundaria = 'ibarra@cdegaulle.cl'
                primaire = 'primaire@cdegaulle.cl'
                cpe = 'asanchez@cdegaulle.cl'
                mantencion = 'lcancino@cdegaulle.cl'
                gerente = 'mdiaz@cdegaulle.cl'
                intendance = 'intendance@cdegaulle.cl'
                admin = 'scpa@cdegaulle.cl'
                RRHH = 'ifuentes@cdegaulle.cl'

                template = loader.get_template('edt/email_licencia.html')
                context = RequestContext(request, {'nombre' : licencias.funcionario.nombre,
                                           'apellido' : licencias.funcionario.apellido1,
                                           'rut': licencias.funcionario.rut,
                                           'funcion' : licencias.funcionario.funcion,
                                           'dv':licencias.funcionario.dv,
                                           'cantidad_dias':licencias.cantidad_dias,
                                           'inicio': licencias.inicio,
                                           'fin' : licencias.fin,
                                           'fecha' : licencias.fecha,
                                           'id' : licencias.id,
                                           'reposo' : licencias.reposo, 
                                           })
                html = template.render(context)
                #msg = EmailMultiAlternatives('Ingreso de licencia médica', html, 'scpa@cdegaulle.cl', [RRHH],cc=[admin],bcc=[admin]) # linea de prueba de correo de licencias
                msg = EmailMultiAlternatives('Ingreso de licencia médica', html, 'scpa@cdegaulle.cl', [proviseur],cc=[RRHH, intendance, mantencion, dirgen, secondaire, director_primaria, director_secundaria, gerente, primaire, cpe],bcc=[admin])
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
                
                flag = 'formset valido'
                #return HttpResponse(suma)
                return redirect('/comprobantelicencia/%d/'%(licencia.id))
            else:
                flag = 'formset no es valido'
                return HttpResponse(flag)        


        else:
            formset = LicenciaFormset()
            formset.fields["funcionario"].queryset = Usuario.objects.exclude(id__in=inactivos)
        data = {
                "usuario" : user,
                "form" : formset,
                "usuarios_filtro" : usuarios_filtro,
                "flag" : flag,
        }

        return render_to_response("edt/licencia.html",data,context_instance=RequestContext(request))
    else:
        return redirect("/main")

def comprobantelicencia(request,pk):
    if not estaLogeado(request):
                return redirect("/login")
    
    user = Usuario.objects.get(id=request.session['usuario'])
    licencia = Licencia.objects.get(id=pk)

    data = {
            "licencia": licencia,
            "usuario" : user,
    }

    return render_to_response("edt/comprobantelicencia.html",data,context_instance=RequestContext(request))

def borralicencia(request,pk):
    if not estaLogeado(request):
        return redirect("/login")
    
    user = Usuario.objects.get(id=request.session['usuario'])
    reemplazos = 0


    licencias = Licencia.objects.all().filter(id=pk)

    if ReemplazoLicencia.objects.all().filter(licencia__in=licencias) > 0:
        for licencia in licencias:

            reemplazos = ReemplazoLicencia.objects.all().filter(licencia__in=licencias)


    data = {
        "reemplazos" : reemplazos,
        "licencias" : licencias,
        "usuario" : user,
    }

    return render_to_response("edt/borrarlicencia.html",data,context_instance=RequestContext(request))

def borrarlicencia(request,pk):
    if not estaLogeado(request):
        return redirect("/login")
    
    user = Usuario.objects.get(id=request.session['usuario'])


    borrar = Licencia.objects.filter(id=pk)
    borrar.delete()

    return redirect("/ConLicencia")


def reemplazoslst(request):
    if not estaLogeado(request):
        return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])

    reemplazos = ReemplazoLicencia.objects.all().order_by('-fecha')
    inicio = " "
    fin = " "
    
    if  user.rol.nivel_acceso == 0:
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "start" in request.GET and request.GET.get("start") != "":
                reemplazos = reemplazos.filter(inicio__gte=request.GET.get("start"))
                inicio = request.GET.get("start")
                inicio = parser.parse(inicio)  


            if "end" in request.GET and request.GET.get("end") != "":
                reemplazos = reemplazos.filter(fin__lte=request.GET.get("end"))
                fin = request.GET.get("end")
                fin = parser.parse(fin)  

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                reemplazos = reemplazos.filter(funcionario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        elif "limpiar" in request.GET:
            return redirect("/ConLicencia")

        paginator = Paginator(reemplazos,60)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            reemplazos = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            reemplazos = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,                 
                 "reemplazos": reemplazos,
                 "usuario" : user,
                 "inicio" : inicio,
                 "fin" : fin,
                 "reemplazos_list" : reemplazos.object_list,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "query_string" : request.META["QUERY_STRING"]
                 }


        return render_to_response("edt/reemplazolst.html",data)
        #return HttpResponse("hola")
    else:
        return redirect("/main")

class ReemplazosPDF(PDFTemplateView):
    template_name="edt/reemplazosPDF.html"

    def get_context_data(self,**kwargs):
        context  =super(ReemplazosPDF,self).get_context_data(
            pagesize="letter",
            title="Listado de Reemplazos",
            **kwargs
            )

        user = Usuario.objects.get(id=self.request.session['usuario'])
        inicio = " "
        fin = " "
        
        Reemplazos = ReemplazoLicencia.objects.all()
    
        if  user.rol.nivel_acceso == 0:
            estamento = Estamento.objects.all()
            usuarios_filtro = Usuario.objects.all().filter(estado=1)
            estamento_filtro = Estamento.objects.all()       

            if "filtrar" in self.request.GET:

                if "start" in self.request.GET and self.request.GET.get("start") != "":
                    Reemplazos = Reemplazos.filter(inicio__gte=self.request.GET.get("start"))
                    inicio = self.request.GET.get("start")
                    inicio = parser.parse(inicio)                    
                    

                if "end" in self.request.GET and self.request.GET.get("end") != "":
                    Reemplazos = Reemplazos.filter(fin__lte=self.request.GET.get("end"))
                    fin = self.request.GET.get("end")
                    fin = parser.parse(fin)
                    

                if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                    estamento = self.request.GET.get("estamento")
                    Reemplazos = Reemplazos.filter(licencia__funcionario__estamento=self.request.GET.get("estamento"))
                    for estament in estamento_filtro:
                        estament.estamento_activo = estament.id == int(estamento)


        context = {
                     "estamento":estamento,                 
                     "Reemplazos": Reemplazos,
                     "usuario" : user,
                     "inicio" : inicio,
                     "fin" : fin,
                     "months" : mkmonth_lst(),
                     "usuarios_filtro" : usuarios_filtro,
                     "estamento_filtro" : estamento_filtro,
                     "query_string" : self.request.META["QUERY_STRING"],
                     }
        context['hoy'] =  datetime.now()

        return context

class ReemplazosExcel(TemplateView):
    def get(self,request,*args,**kwargs):

        user = Usuario.objects.get(id=self.request.session['usuario'])
        inicio = " "
        fin = " "        
        reemplazos = ReemplazoLicencia.objects.all()

        if  user.rol.nivel_acceso == 0:
            estamento = Estamento.objects.all()
            usuarios_filtro = Usuario.objects.all().filter(estado=1)
            estamento_filtro = Estamento.objects.all()       

            if "filtrar" in self.request.GET:

                if "start" in self.request.GET and self.request.GET.get("start") != "":
                    reemplazos = reemplazos.filter(inicio__gte=self.request.GET.get("start"))
                    inicio = self.request.GET.get("start")
                    inicio = parser.parse(inicio)                    
                    

                if "end" in self.request.GET and self.request.GET.get("end") != "":
                    reemplazos = reemplazos.filter(fin__lte=self.request.GET.get("end"))
                    fin = self.request.GET.get("end")
                    fin = parser.parse(fin)
                    

                if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                    estamento = self.request.GET.get("estamento")
                    reemplazos = reemplazos.filter(licencia__funcionario__estamento=self.request.GET.get("estamento"))
                    for estament in estamento_filtro:
                        estament.estamento_activo = estament.id == int(estamento)

            hoy = datetime.now()
            #Creamos el libro de trabajo
            wb= Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws['B1'] = 'Reemplazos'
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('B1:E1')

            # Encabezados

            ws['B3'] = '#'
            ws['C3'] = 'Funcionario'
            ws['D3'] = 'Reemplazante'
            ws['E3'] = 'Horas de reemplazo'
            ws['F3'] = 'Horas a pagar'
            ws['G3'] = 'Fecha'
            ws['H3'] = 'Ingresado por'


            cont = 4
            contador = 1

            for reemplazo in reemplazos:
                ws.cell(row=cont,column=2).value = contador
                ws.cell(row=cont,column=3).value = reemplazo.licencia.funcionario.apellido1+" "+reemplazo.licencia.funcionario.apellido1+" "+reemplazo.licencia.funcionario.nombre
                ws.cell(row=cont,column=4).value = reemplazo.reemplazante.apellido1+" "+reemplazo.reemplazante.apellido2+" "+reemplazo.reemplazante.nombre
                ws.cell(row=cont,column=5).value = reemplazo.horasreemplazo
                ws.cell(row=cont,column=6).value = reemplazo.horaspagar
                ws.cell(row=cont,column=7).value = reemplazo.fecha
                ws.cell(row=cont,column=8).value = reemplazo.ingresadopor.apellido1+" "+reemplazo.ingresadopor.nombre

                cont = cont +1
                contador = contador + 1

            nombre_archivo = "reemplazos.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

def ConLicencia(request):
    if not estaLogeado(request):
        return redirect("/login")
    user = Usuario.objects.get(id=request.session['usuario'])

    licencias = Licencia.objects.all().order_by('-fin')
    hoy = datetime.now()
    inicio = " "
    fin = " "
    
    if  user.rol.nivel_acceso == 0:
        estamento = Estamento.objects.all()
        usuarios_filtro = Usuario.objects.all().filter(estado=1)
        estamento_filtro = Estamento.objects.all()

        if "filtrar" in request.GET:           

            if "persona" in request.GET and request.GET.get("persona") != "0":
                persona = request.GET.get("persona")
                for usuario in usuarios_filtro:
                    usuario.usuario_activo = usuario.id == int(persona)
                licencias = Licencia.objects.all().filter(funcionario=persona).order_by('-inicio')


            if "start" in request.GET and request.GET.get("start") != "":
                licencias = licencias.filter(inicio__gte=request.GET.get("start"))
                inicio = request.GET.get("start")
                inicio = parser.parse(inicio)  


            if "end" in request.GET and request.GET.get("end") != "":
                licencias = licencias.filter(fin__lte=request.GET.get("end"))
                fin = request.GET.get("end")
                fin = parser.parse(fin)  

            if "estamento" in request.GET and request.GET.get("estamento") != "0":
                estamento = request.GET.get("estamento")
                licencias = licencias.filter(funcionario__estamento=request.GET.get("estamento"))
                for estament in estamento_filtro:
                    estament.estamento_activo = estament.id == int(estamento)

        elif "limpiar" in request.GET:
            return redirect("/ConLicencia")

        paginator = Paginator(licencias,60)
        
        try: pagina = int(request.GET.get("page",'1'))
        except ValueError: pagina = 1
            
        try:
            licencias = paginator.page(pagina)
        except (InvalidPage, EmptyPage):
            licencias = paginator.page(paginator.num_pages)


        data = {
                 "estamento":estamento,                 
                 "licencias": licencias,
                 "usuario" : user,
                 "inicio" : inicio,
                 "fin" : fin,
                 "hoy" : hoy,
                 "licencias_list" : licencias.object_list,
                 "months" : mkmonth_lst(),
                 "usuarios_filtro" : usuarios_filtro,
                 "estamento_filtro" : estamento_filtro,
                 "query_string" : request.META["QUERY_STRING"]
                 }


        return render_to_response("edt/conlicencialst.html",data)
        #return HttpResponse("hola")
    else:
        return redirect("/main")

class ConLicenciaPDF(PDFTemplateView):
    template_name="edt/conlicenciaPDF.html"

    def get_context_data(self,**kwargs):
        context  =super(ConLicenciaPDF,self).get_context_data(
            pagesize="letter",
            title="Funcionarios con Licencia",
            **kwargs
            )

        user = Usuario.objects.get(id=self.request.session['usuario'])
        inicio = " "
        fin = " "
        
        licencias = Licencia.objects.all()
    
        if  user.rol.nivel_acceso == 0:
            estamento = Estamento.objects.all()
            usuarios_filtro = Usuario.objects.all().filter(estado=1)
            estamento_filtro = Estamento.objects.all()       

            if "filtrar" in self.request.GET:

                if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                    persona = self.request.GET.get("persona")
                    for usuario in usuarios_filtro:
                        usuario.usuario_activo = usuario.id == int(persona)
                    licencias = Licencia.objects.all().filter(funcionario=persona).order_by('-inicio')

                if "start" in self.request.GET and self.request.GET.get("start") != "":
                    licencias = licencias.filter(inicio__gte=self.request.GET.get("start"))
                    inicio = self.request.GET.get("start")
                    inicio = parser.parse(inicio)                    
                    

                if "end" in self.request.GET and self.request.GET.get("end") != "":
                    licencias = licencias.filter(fin__lte=self.request.GET.get("end"))
                    fin = self.request.GET.get("end")
                    fin = parser.parse(fin)
                    

                if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                    estamento = self.request.GET.get("estamento")
                    licencias = licencias.filter(funcionario__estamento=self.request.GET.get("estamento"))
                    for estament in estamento_filtro:
                        estament.estamento_activo = estament.id == int(estamento)


        context = {
                     "estamento":estamento,                 
                     "licencias": licencias,
                     "usuario" : user,
                     "inicio" : inicio,
                     "fin" : fin,
                     "months" : mkmonth_lst(),
                     "usuarios_filtro" : usuarios_filtro,
                     "estamento_filtro" : estamento_filtro,
                     "query_string" : self.request.META["QUERY_STRING"],
                     }
        context['hoy'] =  datetime.now()

        return context

class ConLicenciaExcel(TemplateView):
    def get(self,request,*args,**kwargs):

        user = Usuario.objects.get(id=self.request.session['usuario'])
        inicio = " "
        fin = " "        
        licencias = Licencia.objects.all()

        if  user.rol.nivel_acceso == 0:
            estamento = Estamento.objects.all()
            usuarios_filtro = Usuario.objects.all().filter(estado=1)
            estamento_filtro = Estamento.objects.all()       

            if "filtrar" in self.request.GET:

                if "persona" in self.request.GET and self.request.GET.get("persona") != "0":
                    persona = self.request.GET.get("persona")
                    for usuario in usuarios_filtro:
                        usuario.usuario_activo = usuario.id == int(persona)
                    licencias = Licencia.objects.all().filter(funcionario=persona).order_by('-inicio')

                if "start" in self.request.GET and self.request.GET.get("start") != "":
                    licencias = licencias.filter(inicio__gte=self.request.GET.get("start"))
                    inicio = self.request.GET.get("start")
                    inicio = parser.parse(inicio)                    
                    

                if "end" in self.request.GET and self.request.GET.get("end") != "":
                    licencias = licencias.filter(fin__lte=self.request.GET.get("end"))
                    fin = self.request.GET.get("end")
                    fin = parser.parse(fin)
                    

                if "estamento" in self.request.GET and self.request.GET.get("estamento") != "0":
                    estamento = self.request.GET.get("estamento")
                    licencias = licencias.filter(funcionario__estamento=self.request.GET.get("estamento"))
                    for estament in estamento_filtro:
                        estament.estamento_activo = estament.id == int(estamento)

            hoy = datetime.now()
            #Creamos el libro de trabajo
            wb= Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws['B1'] = 'FUNCIONARIOS CON LICENCIA'
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('B1:E1')

            # Encabezados

            ws['B3'] = '#'
            ws['C3'] = 'Rut'
            ws['D3'] = 'Funcionario'
            ws['E3'] = 'Cargo'
            ws['F3'] = 'Tipo licencia'
            ws['G3'] = 'características del reposo'
            ws['H3'] = 'Médico'
            ws['I3'] = 'Especialidad'
            ws['J3'] = 'Fecha de ingreso'            
            ws['K3'] = 'Inicio'
            ws['L3'] = 'Fin'
            ws['M3'] = 'Duracion'

            cont = 4
            contador = 1

            for licencia in licencias:
                ws.cell(row=cont,column=2).value = contador
                ws.cell(row=cont,column=3).value = licencia.funcionario.rut+"-"+licencia.funcionario.dv 
                ws.cell(row=cont,column=4).value = licencia.funcionario.apellido1+" "+licencia.funcionario.apellido1+" "+licencia.funcionario.nombre
                ws.cell(row=cont,column=5).value = licencia.funcionario.cargo.nombre
                ws.cell(row=cont,column=6).value = licencia.tipo.nombre
                ws.cell(row=cont,column=7).value = licencia.reposo.nombre
                ws.cell(row=cont,column=8).value = licencia.medico
                ws.cell(row=cont,column=9).value =licencia.especialidad.nombre
                ws.cell(row=cont,column=10).value = licencia.fecha
                ws.cell(row=cont,column=10).number_format = 'dd-mm-yyyy' #formato de salida para fecha en celda
                ws.cell(row=cont,column=11).value = licencia.inicio
                ws.cell(row=cont,column=11).number_format = 'dd-mm-yyyy'
                ws.cell(row=cont,column=12).value = licencia.fin
                ws.cell(row=cont,column=12).number_format = 'dd-mm-yyyy'
                ws.cell(row=cont,column=13).value = licencia.cantidad_dias

                cont = cont +1
                contador = contador + 1

            nombre_archivo = "funcionarios_con_licencia.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response





def wsGenero(request):
    if not estaLogeado(request): 
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])

    # 3 formas de generar arrays====================================    
    sexos = []
    listausr = []    
    femenino = len(Usuario.objects.filter(sexo=2).filter(estado=1))
    masculino = len(Usuario.objects.filter(sexo=1).filter(estado=1))
        
    sexos = [['Varones',masculino],['Damas',femenino]]

    for usuario in Usuario.objects.all().filter(estado=1):       
        listausr.append([usuario.nombre,usuario.apellido1])

   #generacion de array desde una queryset
    data = Usuario.objects.values("jefatura__nombre").annotate(cantidad=Count("jefatura__id")).filter(estado=1)
    
    jefaturas = [ [ x["jefatura__nombre"] , x["cantidad"] ] for x in data]    
        

    #return HttpResponse(json.dumps(sexos))
    return render_to_response("edt/genero.html",{"usuario" : usuarioObj,"sexos" : sexos,"usuarios" : json.dumps(listausr),"jefaturas" : json.dumps(jefaturas)})

def wsJefaturas(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])        
   
    jefaturas = []  
    CPE = len(Usuario.objects.filter(jefatura=1).filter(estado=1))
    PrimSecu = len(Usuario.objects.filter(jefatura=2).filter(estado=1))
    Dirgen = len(Usuario.objects.filter(jefatura=3).filter(estado=1))
    Primaria = len(Usuario.objects.filter(jefatura=4).filter(estado=1))
    Secundaria = len(Usuario.objects.filter(jefatura=5).filter(estado=1))
    Gerencia = len(Usuario.objects.filter(jefatura=6).filter(estado=1))
    Mantencion = len(Usuario.objects.filter(jefatura=7).filter(estado=1))
 
    jefaturas = [['CPE',CPE],['D. Primaria/D. Secundaria',PrimSecu ],['Dir. General', Dirgen],['Primaria',Primaria],['Secundaria',Secundaria],['Gerencia',Gerencia],['Mantencion',Mantencion]]
    # #generacion de array desde una queryset
    # data = Usuario.objects.values("jefatura__nombre").annotate(cantidad=Count("jefatura"))   
    # jefaturas = [ [ x["jefatura__nombre"] , x["cantidad"] ] for x in data]    
        

    #return HttpResponse(json.dumps(jefaturas))
    return render_to_response("edt/jefaturas.html",{"usuario" : usuarioObj,"jefaturas" : jefaturas })

def wsEdades(request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])        
   
    hoy = datetime.now()
    
    data = Usuario.objects.values("fecha_nac").filter(estado=1) #seleccion de fechas de nacimiento  
    edad = [  hoy.year - x["fecha_nac"].year  for x in data] #calculo de edad de funcionario

    edades = [ [""+str(y)+"",edad.count(y)] for y in set(edad)] #generacion de array [edad,cantidad]

    edades = sorted(edades, key=lambda k: k[1], reverse=True) 

    #return HttpResponse(edades)
    return render_to_response("edt/edades.html",{"usuario" : usuarioObj,"edades":edades})

def GraficoPermisos (request):
    if not estaLogeado(request):
        return redirect("/login")
    usuarioObj = Usuario.objects.get(id=request.session['usuario'])
    hoy = datetime.now()
    this_year = hoy.year

    permisos = Permiso.objects.filter(fecha_creacion__year=this_year).exclude(usuario__id=2)
    usuarios = [x.usuario for x in permisos]
    data = [[ "{0}".format(y) , usuarios.count(y)] for y in set(usuarios)]
    data = sorted(data,key=lambda k: k[1])


    #return HttpResponse(data)
    return render_to_response("edt/gpermisos.html",{"usuario": usuarioObj,"gpermisos":data},context_instance=RequestContext(request))